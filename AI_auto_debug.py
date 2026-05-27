# ==============================================
# 阿里云百炼 · PyCharm 智能自动读码调试助手
# 功能：自动读取代码 + 自动调试 + 自动优化
# 无插件 | 不超时 | 免费额度 | 智能分析
# ==============================================
from openai import OpenAI
import tkinter as tk
from tkinter import scrolledtext, ttk
import sys

# ===================== 仅需修改这里 =====================
API_KEY = "sk-4491f6b0f5a7477da6134f1bc909dd38"
# =======================================================

# 固定配置（国内直连，不超时）
BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
MODEL = "qwen-turbo"

# 初始化AI客户端
client = OpenAI(api_key=API_KEY, base_url=BASE_URL)


# ==============================================
# 核心功能：自动读取当前文件的所有代码
# ==============================================
def get_my_code():
    try:
        with open(sys.argv[0], "r", encoding="utf-8") as f:
            lines = f.readlines()
        # 截取用户写的代码（过滤掉AI助手代码）
        code_start = 0
        for i, line in enumerate(lines):
            if "# ========== 下面写你要调试的代码 ==========" in line:
                code_start = i + 1
                break
        return "".join(lines[code_start:]).strip()
    except:
        return "未找到代码，请在指定区域编写代码"


# ==============================================
# AI 自动分析代码（调试/优化/解释）
# ==============================================
def ai_analyze_code():
    code = get_my_code()
    if not code:
        return "未检测到你编写的代码，请在下方区域写入代码~"

    prompt = f"""
    你是专业Python高级工程师，现在自动分析用户编写的代码，完成：
    1. 检查所有Bug、报错、逻辑错误
    2. 给出修复后的完整代码
    3. 优化代码结构、性能、可读性
    4. 解释代码逻辑

    待分析代码：
    {code}
    """

    try:
        res = client.chat.completions.create(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}]
        )
        return res.choices[0].message.content
    except Exception as e:
        return f"AI分析失败：{str(e)}"


# ==============================================
# 智能窗口界面
# ==============================================
class SmartAIAssistant:
    def __init__(self, root):
        self.root = root
        self.root.title("🔥 阿里云百炼 · PyCharm 智能自动调试助手")
        self.root.geometry("850x650")

        # 结果展示区
        self.result_box = scrolledtext.ScrolledText(root, font=("微软雅黑", 11), wrap=tk.WORD, state=tk.DISABLED)
        self.result_box.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

        # 按钮区
        btn_frame = ttk.Frame(root)
        btn_frame.pack(pady=5, padx=10, fill=tk.X)

        self.analyze_btn = ttk.Button(btn_frame, text="🚀 自动分析我的代码", command=self.run_analyze)
        self.analyze_btn.pack(side=tk.LEFT, padx=5)

        self.clear_btn = ttk.Button(btn_frame, text="🧹 清空内容", command=self.clear_result)
        self.clear_btn.pack(side=tk.LEFT, padx=5)

        self.add_msg("系统",
                     "✅ 助手启动成功！\n👉 把你的代码写在文件底部 → 点击【自动分析我的代码】\n👉 AI自动读取、调试、优化你的代码")

    def add_msg(self, title, content):
        self.result_box.config(state=tk.NORMAL)
        self.result_box.insert(tk.END, f"===== {title} =====\n{content}\n\n")
        self.result_box.config(state=tk.DISABLED)
        self.result_box.see(tk.END)

    def run_analyze(self):
        self.add_msg("分析中", "正在自动读取你的代码并调用阿里云百炼分析...")
        result = ai_analyze_code()
        self.add_msg("AI 分析结果", result)

    def clear_result(self):
        self.result_box.config(state=tk.NORMAL)
        self.result_box.delete(1.0, tk.END)
        self.result_box.config(state=tk.DISABLED)


# ========== 下面写你要调试的代码 ==========
# 👇👇👇 把你要调试/优化的代码 写在这里 👇👇👇
import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date
import io
import os

st.set_page_config(
    layout="wide",
    page_title="产品流动性管理系统",
    initial_sidebar_state="collapsed"
)


# ----------------------
# 产品参数加载函数
# ----------------------
def load_product_elements():
    """加载产品要素表"""
    file_path = "产品要素表.xlsx"

    if os.path.exists(file_path):
        try:
            df = pd.read_excel(file_path)
            # 确保所有列都是正确的数据类型
            if "备注" in df.columns:
                df["备注"] = df["备注"].fillna("").astype(str)
            if "申赎渠道" in df.columns:
                df["申赎渠道"] = df["申赎渠道"].fillna("").astype(str)
            if "产品名称" in df.columns:
                df["产品名称"] = df["产品名称"].fillna("").astype(str)

            # 处理简称列（可能有多种命名方式）
            alias_col = None
            for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
                if col_name in df.columns:
                    alias_col = col_name
                    break

            if alias_col:
                df[alias_col] = df[alias_col].fillna("").astype(str)

            # 确保数值列是数值类型
            for col in ["申购到账时间(T+N)", "赎回到账时间(T+N)"]:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            return df
        except Exception as e:
            st.error(f"加载产品要素表失败：{e}")
            return pd.DataFrame()
    else:
        # 如果文件不存在，创建空的DataFrame
        return pd.DataFrame(columns=[
            "产品名称", "申赎渠道", "申购到账时间(T+N)",
            "赎回到账时间(T+N)", "备注"
        ])


def create_product_name_mapping(df_elements):
    """创建产品名称映射：简称 -> 全称"""
    if df_elements.empty or "产品名称" not in df_elements.columns:
        return {}

    mapping = {}

    # 查找简称列
    alias_col = None
    for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
        if col_name in df_elements.columns:
            alias_col = col_name
            break

    if alias_col:
        # 创建简称到全称的映射
        for _, row in df_elements.iterrows():
            full_name = row["产品名称"]
            alias = row[alias_col]
            if alias and str(alias).strip():
                mapping[str(alias).strip()] = full_name
            # 也添加全称到自身的映射
            mapping[full_name] = full_name
    else:
        # 如果没有简称列，只使用产品名称
        for _, row in df_elements.iterrows():
            full_name = row["产品名称"]
            mapping[full_name] = full_name

    return mapping


def save_product_elements(df):
    """保存产品要素表"""
    try:
        df.to_excel("产品要素表.xlsx", index=False)
        return True
    except Exception as e:
        st.error(f"保存失败：{e}")
        return False


# ----------------------
# 主界面：流动性管理表
# ----------------------

# 自定义CSS优化页面布局
st.markdown("""
<style>
    /* 减小标题和文本的间距 */
    .block-container {
        padding-top: 1rem;
        padding-bottom: 0.5rem;
    }

    /* 减小表格行高 */
    .stDataFrame [data-testid="stVerticalBlock"] {
        gap: 0.2rem;
    }

    /* 优化按钮样式 */
    .stButton > button {
        padding: 0.25rem 0.5rem;
        font-size: 0.85rem;
    }

    /* 减小选择框高度 */
    .stSelectbox > div > div {
        min-height: 2.5rem;
    }

    /* 优化表格字体大小 */
    div[data-testid="stDataframe"] {
        font-size: 0.85rem;
    }
</style>
""", unsafe_allow_html=True)

st.title("🏦 产品流动性管理系统")

# ---------- 1. 加载数据 ----------
DEFAULT_FILE = "基金流动性管理总表_test.xlsx"
file_path = DEFAULT_FILE

# 检查文件是否存在
if not os.path.exists(file_path):
    st.error(f"❌ 找不到文件：{file_path}")
    st.stop()


@st.cache_data(ttl=300)  # 缓存5分钟，减少重新加载
def load_all_sheets(file_path):
    """优化版：使用 xls 对象直接读取，避免重复打开文件"""
    # 一次性创建 ExcelFile 对象
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    sheets_data = {}

    # 只读取从第3张sheet开始（索引2）的所有sheet
    target_sheets = xls.sheet_names[2:]  # 从索引2开始，即第3张sheet

    for sheet_name in target_sheets:
        # 使用 xls.parse() 而不是 pd.read_excel()，避免重复读取文件
        df = xls.parse(sheet_name)

        # 取前6列，标准化列名（新增“关联产品”列）
        if df.shape[1] >= 6:
            flow_df = df.iloc[:, :6].copy()
            flow_df.columns = ["日期", "现金流类型", "关联产品", "现金流入", "现金流出", "期末余额"]
        elif df.shape[1] >= 5:
            flow_df = df.iloc[:, :5].copy()
            flow_df.columns = ["日期", "现金流类型", "现金流入", "现金流出", "期末余额"]
            # 插入空的“关联产品”列
            flow_df.insert(2, "关联产品", "")
        else:
            flow_df = df.copy()
            # 补充缺失的列
            for i in range(6 - df.shape[1]):
                flow_df[f"col_{i + df.shape[1] + 1}"] = ""
            flow_df.columns = ["日期", "现金流类型", "关联产品", "现金流入", "现金流出", "期末余额"] + list(
                flow_df.columns[6:])
            flow_df = flow_df.iloc[:, :6]

        # 数值转换 - 使用向量化操作提高性能
        for col in ["现金流入", "现金流出", "期末余额"]:
            flow_df[col] = pd.to_numeric(flow_df[col], errors="coerce")
        # 日期转换 - 指定常见日期格式以避免警告
        flow_df["日期"] = pd.to_datetime(flow_df["日期"], errors="coerce", format="mixed", dayfirst=False)

        # 确保“关联产品”列是字符串类型
        flow_df["关联产品"] = flow_df["关联产品"].fillna("").astype(str)

        sheets_data[sheet_name] = flow_df

    return sheets_data


try:
    sheets_data = load_all_sheets(file_path)
    all_products = list(sheets_data.keys())
except Exception as e:
    st.error(f"读取文件失败：{e}")
    st.stop()

# 检查是否需要重新加载数据（保存后）
if st.session_state.get('reload_data', False):
    # 清除缓存并重新加载
    load_all_sheets.clear()
    sheets_data = load_all_sheets(file_path)
    all_products = list(sheets_data.keys())
    # 清除标志
    st.session_state['reload_data'] = False

# 加载产品要素表并创建名称映射
df_elements = load_product_elements()
name_mapping = create_product_name_mapping(df_elements)

# 创建反向映射：全称 -> 简称（用于显示）
reverse_mapping = {}
if df_elements is not None and not df_elements.empty:
    alias_col = None
    for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
        if col_name in df_elements.columns:
            alias_col = col_name
            break

    if alias_col:
        for _, row in df_elements.iterrows():
            full_name = row["产品名称"]
            alias = row[alias_col]
            if alias and str(alias).strip():
                reverse_mapping[full_name] = str(alias).strip()

# ---------- 2. 产品类型配置（存储在 session_state） ----------
if "product_types" not in st.session_state:
    # 初始默认识别：优先从产品要素表读取，其次根据工作表名称关键词
    st.session_state.product_types = {}

    # 查找类型列（可能有多种命名方式）
    type_col = None
    if not df_elements.empty:
        for col_name in ["产品类型", "类型", "基金类型"]:
            if col_name in df_elements.columns:
                type_col = col_name
                break

    for p in all_products:
        # 1. 优先从产品要素表读取类型
        if type_col and not df_elements.empty:
            # 尝试通过名称匹配
            if p in df_elements["产品名称"].values:
                product_type = df_elements[df_elements["产品名称"] == p].iloc[0][type_col]
                st.session_state.product_types[p] = str(product_type) if pd.notna(product_type) else "子基金"
            # 尝试通过简称匹配
            else:
                alias_col = None
                for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
                    if col_name in df_elements.columns:
                        alias_col = col_name
                        break

                if alias_col and p in df_elements[alias_col].values:
                    product_type = df_elements[df_elements[alias_col] == p].iloc[0][type_col]
                    st.session_state.product_types[p] = str(product_type) if pd.notna(product_type) else "子基金"
                else:
                    # 2. 如果产品要素表没有，根据关键词判断
                    if "母基金" in p or "FOF" in p or "FOF" in p.upper():
                        st.session_state.product_types[p] = "母基金"
                    else:
                        st.session_state.product_types[p] = "子基金"
        else:
            # 3. 如果没有产品要素表，根据关键词判断
            if "母基金" in p or "FOF" in p or "FOF" in p.upper():
                st.session_state.product_types[p] = "母基金"
            else:
                st.session_state.product_types[p] = "子基金"


# ---------- 3. 辅助函数：保存单个 sheet 到 Excel ----------
def save_sheet_to_excel_preserve_rows(sheet_name, df_original, df_updated, wb_path):
    """
    保存数据到Excel，保留原有行结构
    - 新增数据：插入新行
    - 删除数据：清空内容但不删除行
    - 修改数据：更新对应行

    Args:
        sheet_name: 工作表名称
        df_original: 原始DataFrame（用于对比）
        df_updated: 更新后的DataFrame
        wb_path: Excel文件路径
    """
    from openpyxl import load_workbook

    wb = load_workbook(wb_path)
    ws = wb[sheet_name]

    # 获取原始数据的行数（不包括表头）
    original_row_count = len(df_original)
    updated_row_count = len(df_updated)

    # 情况1：更新的行数 <= 原始行数（有删除或不变）
    if updated_row_count <= original_row_count:
        # 清空所有数据行（从第2行开始，第1行是表头）
        for row in range(2, ws.max_row + 1):
            for col in range(1, 7):
                ws.cell(row=row, column=col, value=None)

        # 写入更新后的数据（从第2行开始）
        for r_idx, (_, row_data) in enumerate(df_updated.iterrows(), start=2):
            for c_idx, value in enumerate(row_data[:6], 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # 情况2：更新的行数 > 原始行数（有新增）
    else:
        # 先更新原有的行
        for r_idx, (_, row_data) in enumerate(df_original.iterrows(), start=2):
            for c_idx, value in enumerate(row_data[:6], 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 在末尾插入新行
        new_rows = df_updated.iloc[original_row_count:]
        for _, row_data in new_rows.iterrows():
            ws.append([row_data.iloc[i] if i < len(row_data) else None for i in range(6)])

    wb.save(wb_path)


def sync_to_excel_with_confirmation(selected_product, edited_df, linkage_info=None):
    """
    同步数据到Excel，带确认对话框

    Args:
        selected_product: 产品名称
        edited_df: 编辑后的DataFrame
        linkage_info: 联动信息列表
    """
    # 使用session_state存储待保存的数据
    st.session_state[f'pending_save_{selected_product}'] = {
        'df': edited_df,
        'linkage_info': linkage_info or []
    }

    # 显示确认对话框
    with st.form(key=f'confirm_save_{selected_product}'):
        st.warning("⚠️ 确认将更改同步到Excel文件？")

        if linkage_info:
            st.success(f"🔗 将同时联动 {len(linkage_info)} 个子基金")
            linkage_df = pd.DataFrame(linkage_info)

            # 格式化显示
            styled_df = linkage_df.style.format({
                "金额": "{:,.2f}",
                "当前余额": "{:,.2f}",
                "变动后余额": "{:,.2f}"
                # "比例"列已经是字符串格式（如"50.00%"），不需要额外格式化
            })

            st.dataframe(
                styled_df,
                use_container_width=True,
                hide_index=True,
                height=min(250, len(linkage_info) * 35 + 50)
            )

        col_confirm1, col_confirm2 = st.columns(2)
        with col_confirm1:
            confirm_btn = st.form_submit_button("✅ 确认保存", type="primary", use_container_width=True)
        with col_confirm2:
            cancel_btn = st.form_submit_button("❌ 取消", use_container_width=True)

        if confirm_btn:
            try:
                # 保存到Excel
                original_df = sheets_data[selected_product]
                save_sheet_to_excel_preserve_rows(
                    selected_product,
                    original_df,
                    edited_df,
                    DEFAULT_FILE
                )

                # 如果有联动，也保存到子基金
                if linkage_info:
                    st.info(f"🔗 正在处理 {len(linkage_info)} 个子基金的联动保存...")

                    for linkage in linkage_info:
                        child_product = linkage['子基金']
                        if child_product in sheets_data:
                            try:
                                # 获取子基金的最新数据（可能已经被其他操作修改）
                                child_original = sheets_data[child_product].copy()

                                # 重新计算联动，这次实际添加记录
                                # 从联动信息中提取参数
                                mother_operation_date = pd.to_datetime(linkage['母基金操作日'])
                                amount = linkage['金额']

                                # 判断是申购还是赎回
                                if '申购' in linkage.get('子基金操作', ''):
                                    flow_type = "对子基金申购"
                                else:
                                    flow_type = "对子基金赎回"

                                st.info(f"📝 正在为 {child_product} 添加联动记录：{flow_type}，金额 {amount:,.2f}")

                                # 执行实际的联动操作（preview_mode=False）
                                # 使用单个子基金联动函数
                                handle_single_child_linkage(
                                    selected_product,
                                    mother_operation_date,
                                    child_product,
                                    flow_type,
                                    amount,
                                    preview_mode=False
                                )

                                # 保存更新后的子基金数据
                                child_updated = sheets_data[child_product]
                                save_sheet_to_excel_preserve_rows(
                                    child_product,
                                    child_original,
                                    child_updated,
                                    DEFAULT_FILE
                                )

                                st.success(f"✅ {child_product} 联动记录已保存")
                            except Exception as e:
                                st.error(f"❌ 保存 {child_product} 联动记录失败：{e}")
                                import traceback
                                st.error(traceback.format_exc())

                # 清除待保存状态
                if f'pending_save_{selected_product}' in st.session_state:
                    del st.session_state[f'pending_save_{selected_product}']

                st.success("✅ 保存成功！母子基金数据已同步更新。")

                # 设置标志，触发重新加载数据
                st.session_state['reload_data'] = True

                st.cache_data.clear()
                st.rerun()

            except Exception as e:
                st.error(f"❌ 保存失败：{e}")
                import traceback
                st.error(traceback.format_exc())

        if cancel_btn:
            # 清除待保存状态
            if f'pending_save_{selected_product}' in st.session_state:
                del st.session_state[f'pending_save_{selected_product}']
            st.info("已取消保存")
            st.rerun()


def add_cashflow_record(product_name, date_val, flow_type, related_product, inflow, outflow, balance_after):
    """向指定产品的现金流表追加一行，并返回新的 DataFrame"""
    df = sheets_data[product_name].copy()
    new_row = pd.DataFrame({
        "日期": [date_val],
        "现金流类型": [flow_type],
        "关联产品": [related_product],
        "现金流入": [inflow],
        "现金流出": [outflow],
        "期末余额": [balance_after]
    })
    df = pd.concat([df, new_row], ignore_index=True)
    # 确保数值类型
    for col in ["现金流入", "现金流出", "期末余额"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    sheets_data[product_name] = df
    return df


def get_mother_fund_holdings(mother_fund_name):
    """
    获取母基金持有的子基金份额信息

    Args:
        mother_fund_name: 母基金名称

    Returns:
        dict: {子基金名称: 持有份额}
    """
    holdings = {}

    if mother_fund_name not in sheets_data:
        return holdings

    df = sheets_data[mother_fund_name]

    # 查找所有关联产品不为空的记录
    related_records = df[df["关联产品"].notna() & (df["关联产品"] != "")].copy()

    if related_records.empty:
        return holdings

    # 按关联产品分组，获取每个子基金的最新持仓份额
    # 首先按日期排序，确保获取最新的记录
    related_records = related_records.sort_values(by="日期", ascending=False)

    # 对每个子基金，只保留最新的记录
    seen_products = set()
    for _, row in related_records.iterrows():
        child_product = str(row["关联产品"]).strip()
        if child_product and child_product in all_products:
            # 如果还没处理过这个子基金，使用其最新余额
            if child_product not in seen_products:
                balance = float(row["期末余额"]) if pd.notna(row["期末余额"]) else 0
                holdings[child_product] = balance
                seen_products.add(child_product)

    return holdings


def get_child_product_arrival_days(child_product_name):
    """
    获取子基金的申购/赎回到账时间(T+N)

    Args:
        child_product_name: 子基金名称

    Returns:
        tuple: (申购到账天数, 赎回到账天数)
    """
    subscribe_days = 0
    redeem_days = 0

    # 从产品要素表中查找
    if not df_elements.empty:
        # 尝试通过产品名称匹配
        if child_product_name in df_elements["产品名称"].values:
            product_info = df_elements[df_elements["产品名称"] == child_product_name].iloc[0]
            subscribe_days = int(product_info.get("申购到账时间(T+N)", 0))
            redeem_days = int(product_info.get("赎回到账时间(T+N)", 0))
        else:
            # 尝试通过简称匹配
            alias_col = None
            for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
                if col_name in df_elements.columns:
                    alias_col = col_name
                    break

            if alias_col and child_product_name in df_elements[alias_col].values:
                product_info = df_elements[df_elements[alias_col] == child_product_name].iloc[0]
                subscribe_days = int(product_info.get("申购到账时间(T+N)", 0))
                redeem_days = int(product_info.get("赎回到账时间(T+N)", 0))

    return subscribe_days, redeem_days


def handle_single_child_linkage(mother_fund_name, date_val, child_product_name, flow_type, amount, preview_mode=True):
    """
    处理母基金对单个指定子基金的申购/赎回联动

    联动规则：
    - 母基金申购子基金：母基金流出，指定子基金在T+申购到账日流入
    - 母基金赎回子基金：母基金流入，指定子基金在T+赎回到账日流出

    Args:
        mother_fund_name: 母基金名称
        date_val: 交易日期（母基金的操作日期）
        child_product_name: 指定的子基金名称
        flow_type: 现金流类型（申购/赎回）
        amount: 金额（正数）
        preview_mode: 预览模式，True时只计算不实际添加记录

    Returns:
        list: 生成的联动记录列表（只有一个子基金）
    """
    linkage_records = []

    # 确保 date_val 是 datetime 对象
    if isinstance(date_val, str):
        trade_date = pd.to_datetime(date_val)
    elif isinstance(date_val, pd.Timestamp):
        trade_date = date_val
    else:
        trade_date = pd.to_datetime(date_val)

    # 获取子基金的到账时间
    subscribe_days, redeem_days = get_child_product_arrival_days(child_product_name)

    # 根据操作类型确定现金流方向和到账日期
    if "申购" in flow_type or "对子基金申购" in flow_type:
        # 母基金申购子基金：
        # - 母基金：现金流出（投资出去）
        # - 子基金：在 T-申购到账日 现金流入（注意是T-而不是T+）
        arrival_date = trade_date - pd.Timedelta(days=subscribe_days)

        child_inflow = amount
        child_outflow = 0
        child_flow_type = f"母基金申购到账(T-{subscribe_days})"

    elif "赎回" in flow_type or "对子基金赎回" in flow_type:
        # 母基金赎回子基金：
        # - 母基金：现金流入（收回投资）
        # - 子基金：在 T-赎回到账日 现金流出（注意是T-而不是T+）
        arrival_date = trade_date - pd.Timedelta(days=redeem_days)

        child_inflow = 0
        child_outflow = amount
        child_flow_type = f"母基金赎回到账(T-{redeem_days})"
    else:
        return linkage_records

    # 获取子基金当前余额
    if child_product_name in sheets_data:
        child_df = sheets_data[child_product_name]
        if not child_df.empty:
            current_balance = float(child_df["期末余额"].iloc[-1]) if pd.notna(child_df["期末余额"].iloc[-1]) else 0
        else:
            current_balance = 0
    else:
        current_balance = 0

    # 计算新余额
    new_balance = current_balance + child_inflow - child_outflow

    # 如果不是预览模式，才实际添加记录到子基金
    if not preview_mode:
        add_cashflow_record(
            product_name=child_product_name,
            date_val=arrival_date,
            flow_type=child_flow_type,
            related_product=mother_fund_name,  # 关联回母基金
            inflow=child_inflow,
            outflow=child_outflow,
            balance_after=new_balance
        )

    linkage_records.append({
        "子基金": child_product_name,
        "母基金": mother_fund_name,  # 添加母基金名称
        "母基金操作日": trade_date.strftime("%Y-%m-%d"),
        "到账日期": arrival_date.strftime("%Y-%m-%d"),
        "到账天数": subscribe_days if "申购" in flow_type else redeem_days,
        "比例": "100.00%",  # 单一子基金，100%
        "金额": amount,
        "子基金操作": child_flow_type,
        "当前余额": current_balance,
        "变动后余额": new_balance
    })

    return linkage_records


def handle_mother_fund_cashflow_linkage(mother_fund_name, date_val, flow_type, amount, preview_mode=True):
    """
    处理母基金申购/赎回时的子基金联动

    联动规则：
    - 母基金申购子基金：母基金流出，子基金在T+申购到账日流入
    - 母基金赎回子基金：母基金流入，子基金在T+赎回到账日流出

    Args:
        mother_fund_name: 母基金名称
        date_val: 交易日期（母基金的操作日期）
        flow_type: 现金流类型（申购/赎回）
        amount: 金额（正数）
        preview_mode: 预览模式，True时只计算不实际添加记录

    Returns:
        list: 生成的联动记录列表
    """
    linkage_records = []

    # 获取母基金持有的子基金份额
    holdings = get_mother_fund_holdings(mother_fund_name)

    if not holdings:
        return linkage_records

    # 计算总持有份额
    total_holdings = sum(holdings.values())

    if total_holdings == 0:
        return linkage_records

    # 确保 date_val 是 datetime 对象
    if isinstance(date_val, str):
        trade_date = pd.to_datetime(date_val)
    elif isinstance(date_val, pd.Timestamp):
        trade_date = date_val
    else:
        trade_date = pd.to_datetime(date_val)

    # 按持有比例分配资金到各子基金
    for child_product, child_shares in holdings.items():
        # 计算比例
        ratio = child_shares / total_holdings

        # 按比例分配金额
        linked_amount = amount * ratio

        # 获取子基金的到账时间
        subscribe_days, redeem_days = get_child_product_arrival_days(child_product)

        # 根据操作类型确定现金流方向和到账日期
        if "申购" in flow_type or "对子基金申购" in flow_type:
            # 母基金申购子基金：
            # - 母基金：现金流出（投资出去）
            # - 子基金：在 T+申购到账日 现金流入
            arrival_date = trade_date + pd.Timedelta(days=subscribe_days)

            child_inflow = linked_amount
            child_outflow = 0
            child_flow_type = f"母基金申购到账(T+{subscribe_days})"

        elif "赎回" in flow_type or "对子基金赎回" in flow_type:
            # 母基金赎回子基金：
            # - 母基金：现金流入（收回投资）
            # - 子基金：在 T+赎回到账日 现金流出
            arrival_date = trade_date + pd.Timedelta(days=redeem_days)

            child_inflow = 0
            child_outflow = linked_amount
            child_flow_type = f"母基金赎回到账(T+{redeem_days})"
        else:
            continue

        # 获取子基金当前余额
        if child_product in sheets_data:
            child_df = sheets_data[child_product]
            if not child_df.empty:
                current_balance = float(child_df["期末余额"].iloc[-1]) if pd.notna(child_df["期末余额"].iloc[-1]) else 0
            else:
                current_balance = 0
        else:
            current_balance = 0

        # 计算新余额
        new_balance = current_balance + child_inflow - child_outflow

        # 如果不是预览模式，才实际添加记录到子基金
        if not preview_mode:
            add_cashflow_record(
                product_name=child_product,
                date_val=arrival_date,
                flow_type=child_flow_type,
                related_product=mother_fund_name,  # 关联回母基金
                inflow=child_inflow,
                outflow=child_outflow,
                balance_after=new_balance
            )

        linkage_records.append({
            "子基金": child_product,
            "母基金": mother_fund_name,  # 添加母基金名称
            "母基金操作日": trade_date.strftime("%Y-%m-%d"),
            "到账日期": arrival_date.strftime("%Y-%m-%d"),
            "到账天数": subscribe_days if "申购" in flow_type else redeem_days,
            "比例": f"{ratio:.2%}",
            "金额": linked_amount,
            "子基金操作": child_flow_type,
            "当前余额": current_balance,
            "变动后余额": new_balance
        })

    return linkage_records


# ---------- 4. 主界面：多产品选择与展示 ----------
# st.header("📋 产品现金流明细（可编辑）")

# 多产品选择器 - 更紧凑的布局
st.markdown("**选择要查看的产品（最多3个）**")
col_select1, col_select2, col_select3 = st.columns(3)

with col_select1:
    selected_product1 = st.selectbox("产品 1", [""] + all_products, key="product_select_1",
                                     label_visibility="collapsed")
with col_select2:
    selected_product2 = st.selectbox("产品 2", [""] + all_products, key="product_select_2",
                                     label_visibility="collapsed")
with col_select3:
    selected_product3 = st.selectbox("产品 3", [""] + all_products, key="product_select_3",
                                     label_visibility="collapsed")

# 收集已选择的产品
selected_products = [p for p in [selected_product1, selected_product2, selected_product3] if p]

if not selected_products:
    st.info("👈 请从上方下拉框中选择要查看的产品")
else:
    # 检查是否点击了某个产品进入参数页面
    if st.session_state.get('viewing_product_params', None):
        viewing_product = st.session_state['viewing_product_params']

        # 显示产品参数页面
        st.markdown(f"# 📋 {viewing_product} - 产品参数")

        # 返回按钮
        if st.button("⬅️ 返回现金流管理", type="primary"):
            st.session_state['viewing_product_params'] = None
            st.rerun()

        st.markdown("---")

        # 加载产品要素表
        df_elements = load_product_elements()

        # 使用简称映射查找产品
        product_row = None

        if not df_elements.empty:
            # 查找简称列
            alias_col = None
            for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
                if col_name in df_elements.columns:
                    alias_col = col_name
                    break

            # 1. 首先尝试精确匹配产品名称
            if viewing_product in df_elements["产品名称"].values:
                product_row = df_elements[df_elements["产品名称"] == viewing_product].index[0]
            # 2. 如果有简称列，尝试通过简称匹配
            elif alias_col and viewing_product in df_elements[alias_col].values:
                product_row = df_elements[df_elements[alias_col] == viewing_product].index[0]
            # 3. 尝试通过映射查找全称
            elif name_mapping and viewing_product in name_mapping:
                full_name = name_mapping[viewing_product]
                if full_name in df_elements["产品名称"].values:
                    product_row = df_elements[df_elements["产品名称"] == full_name].index[0]

        if product_row is not None:
            product_info = df_elements.loc[product_row]

            # 显示并编辑产品参数
            col_p1, col_p2 = st.columns(2)
            with col_p1:
                new_channel = st.text_input("申赎渠道", value=str(product_info.get("申赎渠道", "")),
                                            key=f"edit_channel_{viewing_product}")
                new_purchase_days = st.number_input("申购到账时间(T+N)", min_value=0, max_value=30,
                                                    value=int(product_info.get("申购到账时间(T+N)", 0)),
                                                    key=f"edit_purchase_{viewing_product}")
            with col_p2:
                new_redeem_days = st.number_input("赎回到账时间(T+N)", min_value=0, max_value=30,
                                                  value=int(product_info.get("赎回到账时间(T+N)", 0)),
                                                  key=f"edit_redeem_{viewing_product}")
                new_notes = st.text_input("备注", value=str(product_info.get("备注", "")),
                                          key=f"edit_notes_{viewing_product}")

            col_save1, col_save2 = st.columns([1, 3])
            with col_save1:
                if st.button("💾 保存参数", type="primary", key=f"save_product_params_{viewing_product}"):
                    # 更新数据
                    df_elements.loc[product_row, "申赎渠道"] = new_channel
                    df_elements.loc[product_row, "申购到账时间(T+N)"] = new_purchase_days
                    df_elements.loc[product_row, "赎回到账时间(T+N)"] = new_redeem_days
                    df_elements.loc[product_row, "备注"] = new_notes

                    if save_product_elements(df_elements):
                        st.success("✅ 产品参数保存成功！")
                        st.rerun()
            with col_save2:
                if st.button("🔄 取消编辑", key=f"cancel_edit_{viewing_product}"):
                    st.session_state['viewing_product_params'] = None
                    st.rerun()
        else:
            st.info(f"⚠️ 未找到「{viewing_product}」的参数信息，请在产品要素表中添加")

            # 提供快速添加功能
            col_add1, col_add2 = st.columns(2)
            with col_add1:
                add_channel = st.text_input("申赎渠道", placeholder="如：直销、代销",
                                            key=f"add_channel_{viewing_product}")
                add_purchase_days = st.number_input("申购到账时间(T+N)", min_value=0, max_value=30, value=0,
                                                    key=f"add_purchase_{viewing_product}")
            with col_add2:
                add_redeem_days = st.number_input("赎回到账时间(T+N)", min_value=0, max_value=30, value=0,
                                                  key=f"add_redeem_{viewing_product}")
                add_notes = st.text_input("备注", key=f"add_notes_{viewing_product}")

            if st.button("➕ 添加产品参数", type="primary", key=f"add_product_params_{viewing_product}"):
                new_row = pd.DataFrame({
                    "产品名称": [viewing_product],
                    "申赎渠道": [add_channel],
                    "申购到账时间(T+N)": [add_purchase_days],
                    "赎回到账时间(T+N)": [add_redeem_days],
                    "备注": [add_notes]
                })
                df_elements = pd.concat([df_elements, new_row], ignore_index=True)
                if save_product_elements(df_elements):
                    st.success("✅ 产品参数添加成功！")
                    st.rerun()

    else:
        # 创建三列布局，并排显示选中的产品
        product_columns = st.columns(len(selected_products))

        # 添加重置按钮
        col_reset1, col_reset2 = st.columns([1, 4])
        with col_reset1:
            if st.button("🔄 重置所有数据", key="reset_all_data", use_container_width=True):
                # 清除所有产品的 session_state 缓存
                for product in selected_products:
                    edited_session_key = f"edited_data_{product}"
                    original_session_key = f"original_data_{product}"
                    if edited_session_key in st.session_state:
                        del st.session_state[edited_session_key]
                    if original_session_key in st.session_state:
                        del st.session_state[original_session_key]
                st.success("✅ 数据已重置，将从 Excel 重新加载")
                st.rerun()

        with col_reset2:
            st.info("💡 如果数据有异常，请点击「重置所有数据」按钮清除缓存")

        # 先处理所有产品的数据，存储到临时字典中
        product_data_cache = {}

        # 为每个选中的产品显示数据
        for idx, selected_product in enumerate(selected_products):
            with product_columns[idx]:
                # 显示产品类型和可点击的产品名称
                product_type = st.session_state.product_types[selected_product]

                # 使用链接样式的按钮作为产品名称
                if st.button(f"{idx + 1}. {selected_product}", key=f"product_link_{selected_product}",
                             use_container_width=True):
                    st.session_state['viewing_product_params'] = selected_product
                    st.rerun()

                st.caption(f"类型：**{product_type}** | 💡 点击名称查看参数")

                # 显示现金流数据
                # 优先使用 session_state 中的编辑后数据（如果存在）
                edited_session_key = f"edited_data_{selected_product}"
                original_session_key = f"original_data_{selected_product}"
                base_session_key = f"base_data_{selected_product}"  # 基础数据（不含联动记录）

                if edited_session_key in st.session_state and st.session_state[edited_session_key] is not None:
                    # 使用之前编辑后的数据，但清除旧的 _is_new 标记
                    print(f"DEBUG [{selected_product}]: 从 session_state 加载数据")
                    print(
                        f"DEBUG [{selected_product}]: session_state中的行数: {len(st.session_state[edited_session_key])}")
                    df = st.session_state[edited_session_key].copy()

                    # 关键：确保日期列是datetime类型（AgGrid返回的数据可能是字符串）
                    df["日期"] = pd.to_datetime(df["日期"], errors='coerce')

                    print(f"DEBUG [{selected_product}]: 加载后 df 的前3行日期:")
                    print(df['日期'].head(3))
                    print(f"DEBUG [{selected_product}]: 加载后 df 的日期范围: {df['日期'].min()} ~ {df['日期'].max()}")
                    if '_is_new' in df.columns:
                        print(f"DEBUG [{selected_product}]: 检测到 _is_new 列，有 {df['_is_new'].sum()} 个新行")
                        df = df.drop(columns=['_is_new'])  # 清除旧标记，重新检测

                    # 同时从 session_state 加载原始数据（用于对比）
                    if original_session_key in st.session_state:
                        original_df = st.session_state[original_session_key].copy()
                    else:
                        # 如果没有缓存的原始数据，从 Excel 重新加载并处理
                        original_df = sheets_data[selected_product].copy()
                        cutoff_date = pd.Timestamp.today() - timedelta(days=5)
                        original_df["日期"] = pd.to_datetime(original_df["日期"], errors='coerce')

                        # 使用与df相同的过滤逻辑
                        orig_start_idx = None
                        for idx in range(len(original_df)):
                            if pd.notna(original_df.iloc[idx]["日期"]) and original_df.iloc[idx]["日期"] >= cutoff_date:
                                orig_start_idx = idx
                                break

                        if orig_start_idx is not None:
                            original_df = original_df.iloc[orig_start_idx:].copy()
                        else:
                            original_df = original_df[original_df["日期"].isna()].copy()

                        original_df = original_df.reset_index(drop=True)
                        original_df["现金流入"] = original_df["现金流入"].fillna(0).astype(float).round(2)
                        original_df["现金流出"] = original_df["现金流出"].fillna(0).astype(float).round(2)
                        original_df["期末余额"] = original_df["期末余额"].fillna(0).astype(float).round(2)
                        if reverse_mapping:
                            original_df["关联产品"] = original_df["关联产品"].apply(
                                lambda x: reverse_mapping.get(str(x).strip(), x) if str(x).strip() else ""
                            )
                        # 缓存原始数据
                        st.session_state[original_session_key] = original_df.copy()
                else:
                    # 首次加载，从 Excel 读取原始数据
                    print(f"DEBUG [{selected_product}]: 从 Excel 加载原始数据")
                    df = sheets_data[selected_product].copy()

                    # 打印Excel中的完整原始数据（过滤后）
                    print(f"DEBUG [{selected_product}]: Excel原始数据（过滤前 {len(df)} 行）")

                    # 打印Excel中是否有空日期的行
                    df_temp = df.copy()
                    df_temp["日期"] = pd.to_datetime(df_temp["日期"], errors='coerce')
                    blank_rows = df_temp[df_temp["日期"].isna()]
                    if len(blank_rows) > 0:
                        print(f"DEBUG [{selected_product}]: Excel中有 {len(blank_rows)} 行空日期数据")
                        print(blank_rows.head(10))

                    # 保存原始数据的深拷贝（用于对比检测变更）
                    original_df = df.copy()

                    # 对original_df应用过滤和预处理（使用与df相同的逻辑）
                    from datetime import timedelta

                    cutoff_date = pd.Timestamp.today() - timedelta(days=5)
                    original_df["日期"] = pd.to_datetime(original_df["日期"], errors='coerce')

                    # 找到第一个>=cutoff_date的日期的索引
                    orig_start_idx = None
                    for idx in range(len(original_df)):
                        if pd.notna(original_df.iloc[idx]["日期"]) and original_df.iloc[idx]["日期"] >= cutoff_date:
                            orig_start_idx = idx
                            break

                    # 应用相同的过滤逻辑
                    if orig_start_idx is not None:
                        original_df = original_df.iloc[orig_start_idx:].copy()
                    else:
                        # 如果没找到，保留所有空日期行
                        original_df = original_df[original_df["日期"].isna()].copy()

                    original_df = original_df.reset_index(drop=True)

                    # 对original_df应用与df相同的NaN填充和类型转换
                    original_df["现金流入"] = original_df["现金流入"].fillna(0).astype(float).round(2)
                    original_df["现金流出"] = original_df["现金流出"].fillna(0).astype(float).round(2)
                    original_df["期末余额"] = original_df["期末余额"].fillna(0).astype(float).round(2)

                    # 将关联产品列的全称转换为简称（用于显示）
                    if reverse_mapping:
                        original_df["关联产品"] = original_df["关联产品"].apply(
                            lambda x: reverse_mapping.get(str(x).strip(), x) if str(x).strip() else ""
                        )

                    # 将处理后的原始数据保存到 session_state
                    st.session_state[original_session_key] = original_df.copy()

                    # 打印处理后的original_df
                    print(f"DEBUG [{selected_product}]: 处理后的original_df（前3行）:")
                    print(original_df.head(3)[['日期', '关联产品', '现金流入', '现金流出', '期末余额']])

                # 过滤：从today-5开始，找到第一个>=today-5的日期，然后显示该日期及之后的所有行
                from datetime import timedelta

                cutoff_date = pd.Timestamp.today() - timedelta(days=5)
                print(f"DEBUG [{selected_product}]: 过滤前 df行数={len(df)}, cutoff_date={cutoff_date}")

                # 确保日期列是datetime类型
                df["日期"] = pd.to_datetime(df["日期"], errors='coerce')

                # 从today-5开始，找到第一个>=cutoff_date的日期的索引
                start_idx = None
                for idx in range(len(df)):
                    if pd.notna(df.iloc[idx]["日期"]) and df.iloc[idx]["日期"] >= cutoff_date:
                        start_idx = idx
                        break

                # 如果找到起始日期，显示从该日期开始的所有行（包括空白行）
                if start_idx is not None:
                    df = df.iloc[start_idx:].copy()
                    print(f"DEBUG [{selected_product}]: 找到起始日期索引={start_idx}，过滤后 df行数={len(df)}")
                else:
                    # 如果没找到，保留所有空日期行
                    df = df[df["日期"].isna()].copy()
                    print(f"DEBUG [{selected_product}]: 未找到>=cutoff_date的日期，保留所有空日期行，df行数={len(df)}")

                # 重置索引
                df = df.reset_index(drop=True)

                # 打印过滤后的元康1号数据（前10行）
                if selected_product == "元康1号":
                    print(f"DEBUG [元康1号过滤后]: 共{len(df)}行")
                    print(f"DEBUG [元康1号过滤后]: 前10行数据（日期、现金流类型、关联产品）:")
                    for i in range(min(10, len(df))):
                        row = df.iloc[i]
                        print(
                            f"  行{i}: 日期={row['日期']}, 现金流类型={row.get('现金流类型', '')}, 关联产品={row.get('关联产品', '')}")

                # 填充NaN
                df["现金流入"] = df["现金流入"].fillna(0).astype(float).round(2)
                df["现金流出"] = df["现金流出"].fillna(0).astype(float).round(2)
                df["期末余额"] = df["期末余额"].fillna(0).astype(float).round(2)

                # 将关联产品列的全称转换为简称（用于显示）
                if reverse_mapping:
                    df["关联产品"] = df["关联产品"].apply(
                        lambda x: reverse_mapping.get(str(x).strip(), x) if str(x).strip() else ""
                    )

                # 创建简称列表用于下拉选择
                display_options = [""]  # 空选项
                for product in all_products:
                    # 如果有简称显示简称，否则显示全称
                    display_name = reverse_mapping.get(product, product)
                    display_options.append(display_name)

                # 将处理后的df缓存，供后续产品使用
                product_data_cache[selected_product] = df.copy()

                # 先检测联动信息（在渲染表格之前）
                linkage_info = []

                # 检测母基金联动
                if st.session_state.product_types.get(selected_product) == "母基金":
                    # original_df 已经在上面获取，这里是原始数据的深拷贝

                    # 注意：不再排序，完全保持Excel原始顺序
                    # 只需要确保日期列是datetime类型
                    if "日期" in df.columns:
                        df["日期"] = pd.to_datetime(df["日期"], errors='coerce')
                        original_df["日期"] = pd.to_datetime(original_df["日期"], errors='coerce')

                    # 查找新增的行或修改的行（在排序之后）
                    new_row_indices = []

                    # 调试信息：打印数据行数
                    print(f"DEBUG [{selected_product}]: df行数={len(df)}, original_df行数={len(original_df)}")

                    for idx in range(len(df)):
                        # 检查是否超出原始数据范围（真正的新增行）
                        if idx >= len(original_df):
                            new_row_indices.append(idx)
                            print(f"DEBUG: 行{idx} - 新增（超出原始数据范围）")
                        else:
                            # 检查已存在的行是否有修改
                            orig_row = original_df.iloc[idx]
                            edit_row = df.iloc[idx]

                            orig_related = str(orig_row.get("关联产品", "")).strip()
                            edit_related = str(edit_row.get("关联产品", "")).strip()

                            # 四舍五入到2位小数，避免浮点数精度问题
                            orig_inflow = round(
                                float(orig_row.get("现金流入", 0)) if pd.notna(orig_row.get("现金流入", 0)) else 0, 2)
                            edit_inflow = round(
                                float(edit_row.get("现金流入", 0)) if pd.notna(edit_row.get("现金流入", 0)) else 0, 2)
                            orig_outflow = round(
                                float(orig_row.get("现金流出", 0)) if pd.notna(orig_row.get("现金流出", 0)) else 0, 2)
                            edit_outflow = round(
                                float(edit_row.get("现金流出", 0)) if pd.notna(edit_row.get("现金流出", 0)) else 0, 2)

                            # 忽略期末余额的差异（因为会重新计算）
                            if (orig_related != edit_related or
                                    abs(orig_inflow - edit_inflow) > 0.01 or
                                    abs(orig_outflow - edit_outflow) > 0.01):
                                new_row_indices.append(idx)
                                print(
                                    f"DEBUG: 行{idx} - 修改 | 关联产品: '{orig_related}' vs '{edit_related}' | 流入: {orig_inflow} vs {edit_inflow} | 流出: {orig_outflow} vs {edit_outflow}")

                    if new_row_indices:
                        print(f"DEBUG [{selected_product}]: 总共检测到 {len(new_row_indices)} 笔新增/修改记录")
                    else:
                        print(f"DEBUG [{selected_product}]: 没有检测到变更")

                    # 标记新增/修改的行（用于高亮显示）
                    df["_is_new"] = False
                    for idx in new_row_indices:
                        df.iloc[idx, df.columns.get_loc("_is_new")] = True

                    # 检测联动信息
                    for idx in range(len(df)):
                        row = df.iloc[idx]
                        related_product = str(row.get("关联产品", "")).strip()
                        inflow = float(row.get("现金流入", 0)) if pd.notna(row.get("现金流入", 0)) else 0
                        outflow = float(row.get("现金流出", 0)) if pd.notna(row.get("现金流出", 0)) else 0
                        date_val = row.get("日期")

                        # 检测是否为母基金对子基金的申购/赎回操作
                        if related_product and related_product in all_products:
                            if outflow > 0:
                                linkage_records = handle_single_child_linkage(
                                    selected_product, date_val, related_product, "对子基金申购", outflow,
                                    preview_mode=True
                                )
                                linkage_info.extend(linkage_records)
                            elif inflow > 0:
                                linkage_records = handle_single_child_linkage(
                                    selected_product, date_val, related_product, "对子基金赎回", inflow,
                                    preview_mode=True
                                )
                                linkage_info.extend(linkage_records)

                # 检测子基金联动
                elif st.session_state.product_types.get(selected_product) == "子基金":
                    for other_product in selected_products:
                        if other_product != selected_product and st.session_state.product_types.get(
                                other_product) == "母基金":
                            try:
                                # 使用缓存的数据，而不是session_state中的数据
                                other_edited_df = product_data_cache.get(other_product)

                                if other_edited_df is not None and not other_edited_df.empty:
                                    for idx in range(len(other_edited_df)):
                                        row = other_edited_df.iloc[idx]
                                        related_product = str(row.get("关联产品", "")).strip()

                                        if (related_product == selected_product or
                                                related_product == reverse_mapping.get(selected_product,
                                                                                       selected_product) or
                                                name_mapping.get(related_product) == selected_product):
                                            inflow = float(row.get("现金流入", 0)) if pd.notna(
                                                row.get("现金流入", 0)) else 0
                                            outflow = float(row.get("现金流出", 0)) if pd.notna(
                                                row.get("现金流出", 0)) else 0
                                            date_val = row.get("日期")

                                            if outflow > 0:
                                                linkage_records = handle_single_child_linkage(
                                                    other_product, date_val, selected_product, "对子基金申购", outflow,
                                                    preview_mode=True
                                                )
                                                linkage_info.extend(linkage_records)
                                            elif inflow > 0:
                                                linkage_records = handle_single_child_linkage(
                                                    other_product, date_val, selected_product, "对子基金赎回", inflow,
                                                    preview_mode=True
                                                )
                                                linkage_info.extend(linkage_records)
                            except Exception as e:
                                pass

                # 如果是子基金且有联动信息，将联动记录集成到表格数据中
                if st.session_state.product_types.get(selected_product) == "子基金":
                    # 首先，移除旧的联动记录（通过现金流类型识别）
                    # 联动记录的现金流类型包含“母基金”或“申购到账”或“赎回到账”
                    if len(df) > 0:
                        # 识别哪些行是联动记录
                        is_linkage_mask = df.get("现金流类型", "").str.contains("母基金|申购到账|赎回到账", na=False)
                        if is_linkage_mask.any():
                            # 移除旧的联动记录
                            df = df[~is_linkage_mask].copy()
                            df = df.reset_index(drop=True)

                    # 如果有新的联动信息，生成并添加联动记录
                    if linkage_info:
                        preview_records = []
                        for linkage in linkage_info:
                            arrival_date = pd.to_datetime(linkage['到账日期'])
                            inflow = linkage['金额'] if '申购' in linkage.get('子基金操作', '') else 0
                            outflow = linkage['金额'] if '赎回' in linkage.get('子基金操作', '') else 0

                            preview_records.append({
                                "日期": arrival_date,
                                "现金流类型": linkage['子基金操作'],
                                "关联产品": linkage.get('母基金', ''),
                                "现金流入": inflow,
                                "现金流出": outflow,
                                "期末余额": None,
                                "_is_new": True  # 标记为新增，用于高亮
                            })

                        # 将新的联动记录添加到 df
                        if preview_records:
                            print(f"DEBUG [子基金排序前]: df行数={len(df)}, 空日期行数={df['日期'].isna().sum()}")
                            print(f"DEBUG [子基金排序前]: 前5行日期:\n{df['日期'].head()}")

                            df_with_preview = pd.concat([df, pd.DataFrame(preview_records)], ignore_index=True)
                            df_with_preview = df_with_preview.sort_values(by="日期").reset_index(drop=True)

                            print(
                                f"DEBUG [子基金排序后]: df行数={len(df_with_preview)}, 空日期行数={df_with_preview['日期'].isna().sum()}")
                            print(f"DEBUG [子基金排序后]: 前5行日期:\n{df_with_preview['日期'].head()}")

                            df = df_with_preview
                            df["现金流入"] = df["现金流入"].fillna(0).astype(float)
                            df["现金流出"] = df["现金流出"].fillna(0).astype(float)
                            df["期末余额"] = df["期末余额"].fillna(0).astype(float)
                            df["_is_new"] = df["_is_new"].fillna(False)

                # 对所有产品（母基金和子基金）统一计算余额
                if "期末余额" in df.columns:
                    current_balance = 0.0
                    for idx in range(len(df)):
                        inflow = float(df.iloc[idx]["现金流入"]) if pd.notna(df.iloc[idx]["现金流入"]) else 0
                        outflow = float(df.iloc[idx]["现金流出"]) if pd.notna(df.iloc[idx]["现金流出"]) else 0
                        current_balance = current_balance + inflow - outflow
                        df.iloc[idx, df.columns.get_loc("期末余额")] = current_balance

                # 现在渲染表格（使用已处理好的数据）
                # 如果有新增/修改的记录，显示图例说明
                has_new_records = "_is_new" in df.columns and df["_is_new"].any()
                if has_new_records:
                    new_count = df["_is_new"].sum()
                    st.info(f"💡 提示：表格中有 {new_count} 笔新增或修改的记录（包括联动记录）")

                # 先移除内部标记列用于显示
                df_display = df.copy()
                columns_to_hide = ['_is_new']
                if '_is_preview' in df_display.columns:
                    columns_to_hide.append('_is_preview')
                df_for_display = df_display.drop(columns=[col for col in columns_to_hide if col in df_display.columns])

                # 注意：不再对数字进行四舍五入到个位，保留两位小数
                # 这样可以避免 AgGrid 返回的数据与原始数据不一致
                # if '现金流入' in df_for_display.columns:
                #     df_for_display['现金流入'] = df_for_display['现金流入'].round(0).astype(int)
                # if '现金流出' in df_for_display.columns:
                #     df_for_display['现金流出'] = df_for_display['现金流出'].round(0).astype(int)
                # if '期末余额' in df_for_display.columns:
                #     df_for_display['期末余额'] = df_for_display['期末余额'].round(0).astype(int)

                # 确保日期格式为 yyyy-mm-dd
                if '日期' in df_for_display.columns:
                    df_for_display['日期'] = pd.to_datetime(df_for_display['日期']).dt.strftime('%Y-%m-%d')

                # 使用 AgGrid 实现高亮+可编辑的单个表格
                from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

                # 构建 Grid 配置
                gb = GridOptionsBuilder.from_dataframe(df_for_display)

                # 启用编辑功能
                gb.configure_default_column(editable=True, groupable=True, filter=True, resizable=True)

                # 配置关联产品列为下拉选择
                gb.configure_column("关联产品", editable=True, cellEditor='agSelectCellEditor',
                                    cellEditorParams={'values': display_options})

                # 配置日期列
                gb.configure_column("日期", editable=True)

                # 配置数值列
                gb.configure_column("现金流入", editable=True,
                                    type=['numericColumn', 'numberColumnFilter', 'customNumericFormat'])
                gb.configure_column("现金流出", editable=True,
                                    type=['numericColumn', 'numberColumnFilter', 'customNumericFormat'])
                gb.configure_column("期末余额", editable=False,
                                    type=['numericColumn', 'numberColumnFilter', 'customNumericFormat'])  # 余额自动计算，不可编辑

                # 添加行选择功能
                gb.configure_selection('multiple', use_checkbox=True, groupSelectsChildren=True)

                # 配置分页
                gb.configure_pagination(paginationAutoPageSize=True)

                # 如果有新增记录，添加高亮样式
                if has_new_records and '_is_new' in df_display.columns:
                    # 获取新增行的索引
                    new_indices = df_display[df_display['_is_new'] == True].index.tolist()

                    # 定义行样式函数（JavaScript代码）
                    row_style_jscode = JsCode(
                        f"""
                        function(params) {{
                            var rowIndex = params.node.rowIndex;
                            var newIndices = {new_indices};
                            if (newIndices.includes(rowIndex)) {{
                                return {{
                                    'backgroundColor': '#fff3cd',
                                    'color': '#856404',
                                    'fontWeight': 'bold'
                                }};
                            }}
                        }}
                        """
                    )
                    gb.configure_grid_options(rowStyle=row_style_jscode)

                grid_options = gb.build()

                # 渲染 AgGrid 表格
                grid_response = AgGrid(
                    df_for_display,
                    gridOptions=grid_options,
                    allow_unsafe_jscode=True,
                    enable_enterprise_modules=False,
                    update_mode='MODEL_CHANGED',
                    height=min(600, len(df_for_display) * 35 + 100),
                    width='100%',
                    fit_columns_on_grid_load=True,
                    key=f"aggrid_{selected_product}"
                )

                # 获取编辑后的数据
                edited_df = pd.DataFrame(grid_response['data'])

                # 调试：检查AgGrid返回的数据是否与输入一致
                print(f"DEBUG [{selected_product}]: AgGrid返回的数据类型:")
                print(f"  现金流入: {edited_df['现金流入'].dtype if '现金流入' in edited_df.columns else 'N/A'}")
                print(f"  现金流出: {edited_df['现金流出'].dtype if '现金流出' in edited_df.columns else 'N/A'}")
                print(
                    f"  前3行现金流入: {edited_df['现金流入'].head(3).tolist() if '现金流入' in edited_df.columns else 'N/A'}")

                # 更新缓存数据（仅用于当前渲染周期内的多产品联动）
                product_data_cache[selected_product] = edited_df.copy()

                # 注意：不在这里保存 session_state，避免每次渲染都覆盖
                # session_state 只在用户点击“新增空行”或明确保存时才更新

                # 添加"新增空行"和"在选中行上方插入"按钮
                col1, col2, col3 = st.columns([1, 1, 3])
                with col1:
                    if st.button(" 新增空行", key=f"add_row_{selected_product}", use_container_width=True):
                        # 创建新行数据
                        new_row = {
                            '日期': pd.NaT,  # 使用 NaT 表示空日期
                            '现金流类型': '',
                            '关联产品': '',
                            '现金流入': 0.0,
                            '现金流出': 0.0,
                            '期末余额': 0.0,
                            '_is_new': True
                        }
                        # 将新行添加到DataFrame末尾
                        new_row_df = pd.DataFrame([new_row])
                        edited_df = pd.concat([edited_df, new_row_df], ignore_index=True)
                        print(f"DEBUG [{selected_product}]: 新增行后，edited_df行数: {len(edited_df)}")
                        # 更新session_state
                        st.session_state[f"edited_data_{selected_product}"] = edited_df
                        print(
                            f"DEBUG [{selected_product}]: 已保存到 session_state，行数: {len(st.session_state[f'edited_data_{selected_product}'])}")
                        # 标记已添加新行，避免后续代码覆盖
                        st.session_state[f"row_added_{selected_product}"] = True
                        # 立即刷新页面，让新行生效
                        st.rerun()

                with col2:
                    if st.button(" 在选中行上方插入", key=f"insert_row_{selected_product}", use_container_width=True):
                        # 获取选中的行
                        selected_rows = grid_response.get('selected_rows', None)
                        # 检查是否有选中的行
                        if selected_rows is not None and not selected_rows.empty:
                            # 获取选中行的索引（取第一个）
                            # selected_rows 是 DataFrame，索引可能是字符串，需要转换为整数
                            if len(selected_rows) > 0:
                                try:
                                    selected_index = int(selected_rows.index[0])
                                except (ValueError, TypeError) as e:
                                    st.warning(f"️ 无法解析选中行的索引: {e}")
                                    print(
                                        f"DEBUG [{selected_product}]: selected_rows.index[0] = {selected_rows.index[0]}, type = {type(selected_rows.index[0])}")
                                    print(f"DEBUG [{selected_product}]: selected_rows:\n{selected_rows}")
                                    selected_index = None

                                if selected_index is not None:
                                    # 创建新行数据
                                    new_row = {
                                        '日期': pd.NaT,
                                        '现金流类型': '',
                                        '关联产品': '',
                                        '现金流入': 0.0,
                                        '现金流出': 0.0,
                                        '期末余额': 0.0,
                                        '_is_new': True
                                    }
                                    new_row_df = pd.DataFrame([new_row])

                                    # 在选中行上方插入
                                    top_part = edited_df.iloc[:selected_index]
                                    bottom_part = edited_df.iloc[selected_index:]
                                    edited_df = pd.concat([top_part, new_row_df, bottom_part], ignore_index=True)

                                    print(
                                        f"DEBUG [{selected_product}]: 在行{selected_index}上方插入新行，edited_df行数: {len(edited_df)}")

                                    # 更新session_state
                                    st.session_state[f"edited_data_{selected_product}"] = edited_df
                                    st.session_state[f"row_added_{selected_product}"] = True

                                    # 立即刷新页面，让新插入的行生效
                                    st.rerun()
                                else:
                                    st.warning("️ 无法获取选中行的位置")
                        else:
                            st.warning("️ 请先选中一行（点击行首的复选框）")

                with col3:
                    st.info("💡 点击「新增空行」在末尾添加，或先选中一行再点击「在选中行上方插入」")

                # 存储编辑后的数据到 session_state
                # 关键：无论是否有新增行，都需要重新计算余额
                # 重新计算余额（用户可能修改了现金流数据，或插入了新行）
                if "期末余额" in edited_df.columns:
                    current_balance = 0.0
                    for idx in range(len(edited_df)):
                        inflow = float(edited_df.iloc[idx]["现金流入"]) if pd.notna(
                            edited_df.iloc[idx]["现金流入"]) else 0
                        outflow = float(edited_df.iloc[idx]["现金流出"]) if pd.notna(
                            edited_df.iloc[idx]["现金流出"]) else 0
                        current_balance = current_balance + inflow - outflow
                        edited_df.iloc[idx, edited_df.columns.get_loc("期末余额")] = current_balance
                    print(f"DEBUG [{selected_product}]: 重新计算余额后，最后一行余额={current_balance}")

                # 保存编辑后的数据到 session_state
                st.session_state[f"edited_data_{selected_product}"] = edited_df

                # 清除新增行标记（如果存在）
                if st.session_state.get(f"row_added_{selected_product}", False):
                    st.session_state[f"row_added_{selected_product}"] = False

                # 关键：自动刷新页面，让AgGrid显示重新计算后的余额
                # 注意：如果有新增行操作，不执行自动刷新（由按钮自己的 st.rerun() 处理）
                if not st.session_state.get(f"row_added_{selected_product}", False):
                    # 使用一个标记来避免无限循环刷新
                    if 'last_refresh_time' not in st.session_state:
                        st.session_state['last_refresh_time'] = 0

                    import time

                    current_time = time.time()
                    # 如果距离上次刷新超过1秒，才刷新（避免无限循环）
                    if current_time - st.session_state['last_refresh_time'] > 1.0:
                        st.session_state['last_refresh_time'] = current_time
                        st.rerun()

                # 保存按钮 - 触发确认对话框
                if st.button(f"💾 保存到Excel", type="primary", key=f"save_{selected_product}",
                             use_container_width=True):
                    # 将简称转换回全称（用于保存到Excel）
                    df_for_save = edited_df.copy()
                    if name_mapping:
                        df_for_save["关联产品"] = df_for_save["关联产品"].apply(
                            lambda x: name_mapping.get(str(x).strip(), x) if str(x).strip() else ""
                        )

                    # 调用确认保存函数
                    sync_to_excel_with_confirmation(selected_product, df_for_save, linkage_info)

st.markdown("---")
st.caption("💡 提示：可以同时选择最多3个产品进行对比查看，点击「查看/修改产品参数」按钮可以查看和编辑产品参数")

# ==============================================
# 启动助手
if __name__ == "__main__":
    window = tk.Tk()
    app = SmartAIAssistant(window)
    window.mainloop()
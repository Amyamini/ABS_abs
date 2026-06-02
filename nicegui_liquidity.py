"""
产品流动性管理系统 - NiceGUI 版本
基于 NiceGUI + AG Grid 重写，替代 Streamlit 版本。

运行方式:
    pip install nicegui pandas openpyxl
    python nicegui_liquidity.py

对比 Streamlit 版本的改进:
    1. 无全局重运行 — 事件驱动，状态常驻内存
    2. AG Grid — 原生可编辑表格，支持 Ctrl+Z 撤销、复制粘贴
    3. 普通 Python 对象管理状态 — 不用 st.session_state
    4. 原生 Dialog — 保存确认/参数编辑更自然
    5. 页面对比不抢焦点，不需要 fragment hack
"""

import os
from datetime import datetime, date, timedelta
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from nicegui import ui

# ============================================================
# 常量
# ============================================================
DEFAULT_FILE = "基金流动性管理总表_test.xlsx"
ELEMENTS_FILE = "产品要素表.xlsx"
TARGET_COLS = ["日期", "现金流类型", "关联产品", "现金流入", "现金流出", "期末余额", "备注"]

# ============================================================
# 工具函数
# ============================================================
def _find_alias_column(df: pd.DataFrame) -> Optional[str]:
    for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
        if col_name in df.columns:
            return col_name
    return None


def _fix_excel_date(val):
    if pd.isna(val):
        return val
    if hasattr(val, "strftime"):
        return val
    try:
        num = float(val)
        if 60 < num < 200000:
            return datetime(1899, 12, 30) + timedelta(days=int(num))
    except (ValueError, TypeError):
        pass
    return val


# ============================================================
# 产品要素表 I/O
# ============================================================
def load_product_elements() -> pd.DataFrame:
    if not os.path.exists(ELEMENTS_FILE):
        return pd.DataFrame(columns=["产品名称", "申赎渠道", "申购到账时间(T+N)", "赎回到账时间(T+N)", "备注"])
    try:
        df = pd.read_excel(ELEMENTS_FILE)
        for col in ["备注", "申赎渠道", "产品名称"]:
            if col in df.columns:
                df[col] = df[col].fillna("").astype(str)
        alias_col = _find_alias_column(df)
        if alias_col:
            df[alias_col] = df[alias_col].fillna("").astype(str)
        for col in ["申购到账时间(T+N)", "赎回到账时间(T+N)"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
        return df
    except Exception as e:
        ui.notify(f"加载产品要素表失败：{e}", type="negative")
        return pd.DataFrame()


def save_product_elements(df: pd.DataFrame) -> bool:
    try:
        df.to_excel(ELEMENTS_FILE, index=False)
        return True
    except Exception as e:
        ui.notify(f"保存失败：{e}", type="negative")
        return False


def create_product_name_mapping(df_elements: pd.DataFrame) -> dict:
    if df_elements.empty or "产品名称" not in df_elements.columns:
        return {}
    mapping = {}
    alias_col = _find_alias_column(df_elements)
    if alias_col:
        for _, row in df_elements.iterrows():
            full_name = row["产品名称"]
            alias = row[alias_col]
            if alias and str(alias).strip():
                mapping[str(alias).strip()] = full_name
            mapping[full_name] = full_name
    else:
        for _, row in df_elements.iterrows():
            mapping[row["产品名称"]] = row["产品名称"]
    return mapping


# ============================================================
# Excel Sheet 加载
# ============================================================
def load_all_sheets(file_path: str) -> dict[str, pd.DataFrame]:
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    sheets_data = {}
    sheet_names = xls.sheet_names
    target_sheets = sheet_names[2:-4] if len(sheet_names) > 6 else sheet_names[2:]
    for sheet_name in target_sheets:
        df = xls.parse(sheet_name)
        sheets_data[sheet_name] = _normalize_sheet_df(df)
    return sheets_data


def _normalize_sheet_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.shape[1] >= 7:
        flow_df = df.iloc[:, :7].copy()
        flow_df.columns = TARGET_COLS
    elif df.shape[1] >= 6:
        flow_df = df.iloc[:, :6].copy()
        flow_df.columns = TARGET_COLS[:6]
        flow_df["备注"] = ""
    elif df.shape[1] >= 5:
        flow_df = df.iloc[:, :5].copy()
        flow_df.columns = ["日期", "现金流类型", "现金流入", "现金流出", "期末余额"]
        flow_df.insert(2, "关联产品", "")
        flow_df["备注"] = ""
    else:
        flow_df = df.copy()
        for i in range(7 - df.shape[1]):
            flow_df[f"col_{i + df.shape[1] + 1}"] = ""
        flow_df.columns = TARGET_COLS + list(flow_df.columns[7:])
        flow_df = flow_df.iloc[:, :7]
    for col in ["现金流入", "现金流出", "期末余额"]:
        flow_df[col] = pd.to_numeric(flow_df[col], errors="coerce")
    flow_df["日期"] = flow_df["日期"].apply(_fix_excel_date)
    flow_df["日期"] = pd.to_datetime(flow_df["日期"], errors="coerce")
    for col in ["现金流类型", "关联产品", "备注"]:
        if col in flow_df.columns:
            flow_df[col] = flow_df[col].fillna("").astype(str)
    return flow_df


# ============================================================
# 业务逻辑
# ============================================================
def sort_cashflow_rows_preserve_blank_positions(df: pd.DataFrame) -> pd.DataFrame:
    if "日期" not in df.columns:
        return df
    if not df["日期"].isna().any():
        df["_orig_idx"] = range(len(df))
        return df.sort_values(by=["日期", "_orig_idx"], na_position="last").drop(
            columns=["_orig_idx"]
        ).reset_index(drop=True)
    blank_positions = df[df["日期"].isna()].index.tolist()
    non_blank = df[df["日期"].notna()].copy()
    non_blank["_orig_idx"] = non_blank.index.tolist()
    non_blank = non_blank.sort_values(by=["日期", "_orig_idx"], na_position="last").drop(
        columns=["_orig_idx"]
    ).reset_index(drop=True)
    result_rows = []
    nb_iter = iter(non_blank.to_dict("records"))
    blank_set = set(blank_positions)
    for i in range(len(df)):
        if i in blank_set:
            result_rows.append(df.iloc[i].to_dict())
        else:
            result_rows.append(next(nb_iter))
    return pd.DataFrame(result_rows)


def resolve_cutoff_start(df: pd.DataFrame, cutoff_date) -> Optional[int]:
    for row_i in range(len(df)):
        if pd.notna(df.iloc[row_i]["日期"]) and df.iloc[row_i]["日期"] >= cutoff_date:
            return row_i
    for row_i in range(len(df) - 1, -1, -1):
        if pd.notna(df.iloc[row_i]["日期"]):
            return row_i
    return None


def calc_balance_preserve_first(df: pd.DataFrame) -> pd.DataFrame:
    if "期末余额" not in df.columns or len(df) == 0:
        return df
    first_balance = df["期末余额"].iloc[0]
    if pd.isna(first_balance):
        first_balance = 0.0
    else:
        first_balance = float(first_balance)
    net_flow = df["现金流入"].fillna(0) - df["现金流出"].fillna(0)
    df["期末余额"] = first_balance + net_flow.cumsum() - net_flow.iloc[0]
    df["期末余额"] = df["期末余额"].round(2)
    return df


def get_product_submit_days(product_name: str, df_elements: pd.DataFrame, name_mapping: dict) -> tuple:
    if df_elements.empty or "产品名称" not in df_elements.columns:
        return 0, 0
    if product_name in df_elements["产品名称"].values:
        info = df_elements[df_elements["产品名称"] == product_name].iloc[0]
        return int(info.get("申购到账时间(T+N)", 0)), int(info.get("赎回到账时间(T+N)", 0))
    alias_col = _find_alias_column(df_elements)
    if alias_col and product_name in df_elements[alias_col].values:
        info = df_elements[df_elements[alias_col] == product_name].iloc[0]
        return int(info.get("申购到账时间(T+N)", 0)), int(info.get("赎回到账时间(T+N)", 0))
    if product_name in name_mapping and name_mapping[product_name] in df_elements["产品名称"].values:
        info = df_elements[df_elements["产品名称"] == name_mapping[product_name]].iloc[0]
        return int(info.get("申购到账时间(T+N)", 0)), int(info.get("赎回到账时间(T+N)", 0))
    return 0, 0


def get_child_product_arrival_days(child_product_name: str, df_elements: pd.DataFrame) -> tuple:
    if df_elements.empty:
        return 0, 0
    if child_product_name in df_elements["产品名称"].values:
        info = df_elements[df_elements["产品名称"] == child_product_name].iloc[0]
        return int(info.get("申购到账时间(T+N)", 0)), int(info.get("赎回到账时间(T+N)", 0))
    alias_col = _find_alias_column(df_elements)
    if alias_col and child_product_name in df_elements[alias_col].values:
        info = df_elements[df_elements[alias_col] == child_product_name].iloc[0]
        return int(info.get("申购到账时间(T+N)", 0)), int(info.get("赎回到账时间(T+N)", 0))
    return 0, 0


def get_mother_fund_holdings(mother_fund_name: str, sheets_data: dict, all_products: list) -> dict:
    holdings = {}
    if mother_fund_name not in sheets_data:
        return holdings
    df = sheets_data[mother_fund_name]
    related = df[df["关联产品"].notna() & (df["关联产品"] != "")].copy()
    if related.empty:
        return holdings
    related = related.sort_values(by="日期", ascending=False)
    seen = set()
    for _, row in related.iterrows():
        child = str(row["关联产品"]).strip()
        if child and child in all_products and child not in seen:
            balance = float(row["期末余额"]) if pd.notna(row["期末余额"]) else 0
            holdings[child] = balance
            seen.add(child)
    return holdings


def handle_single_child_linkage(
    mother_fund_name: str, date_val, child_product_name: str, flow_type: str,
    amount: float, sheets_data: dict, df_elements: pd.DataFrame,
    preview_mode: bool = True,
) -> list[dict]:
    trade_date = pd.to_datetime(date_val)
    sub_days, red_days = get_child_product_arrival_days(child_product_name, df_elements)

    if "申购" in flow_type:
        arrival_date = trade_date + pd.offsets.BDay(sub_days)
        child_inflow, child_outflow = amount, 0
        child_flow_type = f"母基金申购到账(T+{sub_days})"
    elif "赎回" in flow_type:
        arrival_date = trade_date + pd.offsets.BDay(red_days)
        child_inflow, child_outflow = 0, amount
        child_flow_type = f"母基金赎回到账(T+{red_days})"
    else:
        return []

    current_balance = 0
    if child_product_name in sheets_data:
        child_df = sheets_data[child_product_name]
        if not child_df.empty and pd.notna(child_df["期末余额"].iloc[-1]):
            current_balance = float(child_df["期末余额"].iloc[-1])

    new_balance = current_balance + child_inflow - child_outflow

    if not preview_mode:
        _add_cashflow_record(
            sheets_data, child_product_name, arrival_date,
            child_flow_type, mother_fund_name, child_inflow, child_outflow, new_balance,
        )

    return [{
        "子基金": child_product_name,
        "母基金": mother_fund_name,
        "母基金操作日": trade_date.strftime("%Y-%m-%d") if pd.notna(trade_date) else "",
        "到账日期": arrival_date.strftime("%Y-%m-%d") if pd.notna(arrival_date) else "",
        "到账天数": sub_days if "申购" in flow_type else red_days,
        "比例": "100.00%",
        "金额": round(amount, 2),
        "子基金操作": child_flow_type,
        "当前余额": round(current_balance, 2),
        "变动后余额": round(new_balance, 2),
    }]


def _add_cashflow_record(
    sheets_data: dict, product_name: str, date_val, flow_type: str,
    related_product: str, inflow: float, outflow: float, balance_after: float,
) -> pd.DataFrame:
    df = sheets_data[product_name].copy()
    new_row = pd.DataFrame([{
        "日期": date_val, "现金流类型": flow_type, "关联产品": related_product,
        "现金流入": inflow, "现金流出": outflow, "期末余额": balance_after, "备注": "",
    }])
    df = pd.concat([df, new_row], ignore_index=True)
    for col in ["现金流入", "现金流出", "期末余额"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    sheets_data[product_name] = df
    return df


def get_upcoming_events(sheets_data: dict, df_elements: pd.DataFrame,
                        name_mapping: dict, n_days: int = 5) -> list[dict]:
    today = datetime.now().date()
    end_date = today + timedelta(days=n_days)
    events = []
    for product_name, df in sheets_data.items():
        if df.empty or "日期" not in df.columns:
            continue
        purchase_days, redeem_days = get_product_submit_days(product_name, df_elements, name_mapping)
        for _, row in df.iterrows():
            row_date = row["日期"]
            if pd.isna(row_date):
                continue
            row_date = row_date.date() if hasattr(row_date, "date") else row_date
            inflow = float(row.get("现金流入", 0)) if pd.notna(row.get("现金流入", 0)) else 0
            outflow = float(row.get("现金流出", 0)) if pd.notna(row.get("现金流出", 0)) else 0
            if inflow == 0 and outflow == 0:
                continue
            if outflow > 0:
                days, action_type, amount = purchase_days, "申购", outflow
            else:
                days, action_type, amount = redeem_days, "赎回", inflow
            ts = pd.Timestamp(row_date)
            submit_date = (ts - pd.offsets.BDay(days)).date() if days > 0 else row_date
            if submit_date < today or submit_date > end_date:
                continue
            events.append({
                "提交日期": submit_date, "到账日期": row_date, "产品": product_name,
                "操作类型": action_type, "金额": amount, "到账天数": days,
            })
    events.sort(key=lambda x: (x["提交日期"], x["产品"]))
    return events


def save_sheet_to_excel(sheet_name: str, df_original: pd.DataFrame,
                        df_updated: pd.DataFrame, wb_path: str):
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]
    orig_count, upd_count = len(df_original), len(df_updated)
    if upd_count <= orig_count:
        for row in range(2, ws.max_row + 1):
            for col in range(1, 8):
                ws.cell(row=row, column=col, value=None)
        for r_idx, (_, row_data) in enumerate(df_updated.iterrows(), start=2):
            for c_idx, value in enumerate(row_data[:7], 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    else:
        for r_idx, (_, row_data) in enumerate(df_original.iterrows(), start=2):
            for c_idx, value in enumerate(row_data[:7], 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        for _, row_data in df_updated.iloc[orig_count:].iterrows():
            ws.append([row_data.iloc[i] if i < len(row_data) else None for i in range(7)])
    wb.save(wb_path)


def classify_products(all_products: list, df_elements: pd.DataFrame) -> dict:
    types = {}
    type_col = None
    if not df_elements.empty:
        for col_name in ["产品类型", "类型", "基金类型"]:
            if col_name in df_elements.columns:
                type_col = col_name
                break
    alias_col = _find_alias_column(df_elements)
    for p in all_products:
        if type_col and not df_elements.empty:
            info = None
            if p in df_elements["产品名称"].values:
                info = df_elements[df_elements["产品名称"] == p].iloc[0]
            elif alias_col and p in df_elements[alias_col].values:
                info = df_elements[df_elements[alias_col] == p].iloc[0]
            if info is not None:
                pt = info[type_col]
                types[p] = str(pt) if pd.notna(pt) else "子基金"
            else:
                types[p] = "母基金" if ("母基金" in p or "FOF" in p.upper()) else "子基金"
        else:
            types[p] = "母基金" if ("母基金" in p or "FOF" in p.upper()) else "子基金"
    return types


# ============================================================
# 应用状态
# ============================================================
class AppState:
    def __init__(self):
        self.sheets_data: dict[str, pd.DataFrame] = {}
        self.df_elements = pd.DataFrame()
        self.name_mapping: dict = {}
        self.reverse_mapping: dict = {}
        self.all_products: list[str] = []
        self.product_types: dict[str, str] = {}
        self.selected_product: Optional[str] = None
        self.current_page = "流动性管理"
        # 编辑缓存: {product: [row_dict, ...]}
        self.edit_cache: dict[str, list[dict]] = {}
        # 原始完整数据
        self.full_original: dict[str, pd.DataFrame] = {}
        # 原始切片数据
        self.original_slice: dict[str, pd.DataFrame] = {}

    def load(self, file_path: str = DEFAULT_FILE) -> bool:
        if not os.path.exists(file_path):
            return False
        self.sheets_data = load_all_sheets(file_path)
        self.df_elements = load_product_elements()
        self.name_mapping = create_product_name_mapping(self.df_elements)
        self.all_products = list(self.sheets_data.keys())
        self._filter_products()
        self.product_types = classify_products(self.all_products, self.df_elements)
        self._build_reverse_mapping()
        self._init_edit_cache()
        if self.all_products:
            self.selected_product = self.all_products[0]
        return True

    def _filter_products(self):
        if self.df_elements.empty or "产品名称" not in self.df_elements.columns:
            return
        valid = set(self.df_elements["产品名称"].dropna().unique())
        self.all_products = [
            p for p in self.all_products
            if p in valid or self.name_mapping.get(p) in valid
        ]

    def _build_reverse_mapping(self):
        self.reverse_mapping = {}
        if self.df_elements.empty:
            return
        alias_col = _find_alias_column(self.df_elements)
        if alias_col:
            for _, row in self.df_elements.iterrows():
                full = row["产品名称"]
                alias = row[alias_col]
                if alias and str(alias).strip():
                    self.reverse_mapping[full] = str(alias).strip()

    def _init_edit_cache(self):
        cutoff = pd.Timestamp.today() - timedelta(days=5)
        for product in self.all_products:
            df = self.sheets_data[product].copy()
            df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
            start = resolve_cutoff_start(df, cutoff)
            if start is not None:
                df = df.iloc[start:].copy()
            else:
                df = df[df["日期"].isna()].copy()
            df = df.reset_index(drop=True)
            for col in ["现金流入", "现金流出", "期末余额"]:
                df[col] = df[col].fillna(0).astype(float).round(2)
            df["关联产品"] = df["关联产品"].apply(
                lambda x: self.reverse_mapping.get(str(x).strip(), x) if str(x).strip() else ""
            )
            df = sort_cashflow_rows_preserve_blank_positions(df)
            calc_balance_preserve_first(df)
            self.full_original[product] = self.sheets_data[product].copy()
            self.original_slice[product] = df.copy()
            self.edit_cache[product] = self._df_to_rows(df)

    def _df_to_rows(self, df: pd.DataFrame) -> list[dict]:
        rows = []
        for _, row in df.iterrows():
            r = {}
            for col in TARGET_COLS:
                val = row.get(col)
                if col == "日期" and pd.notna(val) and hasattr(val, "strftime"):
                    r[col] = val.strftime("%Y-%m-%d")
                elif col in ("现金流入", "现金流出", "期末余额"):
                    r[col] = round(float(val), 2) if pd.notna(val) else 0.0
                else:
                    r[col] = str(val) if pd.notna(val) and str(val) != "nan" else ""
            rows.append(r)
        return rows

    def get_display_name(self, product: str) -> str:
        return self.reverse_mapping.get(product, product)

    def reload(self):
        self.sheets_data = load_all_sheets(DEFAULT_FILE)
        self.all_products = list(self.sheets_data.keys())
        self._filter_products()
        if self.selected_product not in self.all_products and self.all_products:
            self.selected_product = self.all_products[0]
        self._init_edit_cache()

    def get_grid_rows(self, product: str) -> list[dict]:
        return self.edit_cache.get(product, [])

    def set_grid_rows(self, product: str, rows: list[dict]):
        self.edit_cache[product] = rows

    def compute_linkage(self, product: str) -> list[dict]:
        """计算指定产品的联动预览"""
        ptype = self.product_types.get(product, "子基金")
        if ptype != "母基金":
            return self._compute_child_linkage(product)

        df = pd.DataFrame(self.edit_cache.get(product, []))
        if df.empty:
            return []

        original_df = self.original_slice.get(product, pd.DataFrame())
        records = []

        for row_i, row in df.iterrows():
            related = str(row.get("关联产品", "")).strip()
            inflow = float(row.get("现金流入", 0)) if pd.notna(row.get("现金流入", 0)) else 0
            outflow = float(row.get("现金流出", 0)) if pd.notna(row.get("现金流出", 0)) else 0

            if not related:
                continue
            # Resolve display name to full product name
            resolved = related
            if related not in self.all_products:
                for full_name in self.all_products:
                    if self.reverse_mapping.get(full_name, full_name) == related:
                        resolved = full_name
                        break
            if resolved not in self.all_products:
                continue
            if inflow == 0 and outflow == 0:
                continue

            is_new = row_i >= len(original_df)
            if not is_new:
                orig = original_df.iloc[row_i]
                o_rel = str(orig.get("关联产品", "")).strip()
                o_in = float(orig.get("现金流入", 0) or 0)
                o_out = float(orig.get("现金流出", 0) or 0)
                if o_rel == related and abs(o_in - inflow) < 0.01 and abs(o_out - outflow) < 0.01:
                    continue

            ft = "对子基金申购" if outflow > 0 else "对子基金赎回"
            amt = outflow if outflow > 0 else inflow
            records.extend(
                handle_single_child_linkage(
                    product, row.get("日期"), resolved, ft, amt,
                    self.sheets_data, self.df_elements, preview_mode=True,
                )
            )

        return records

    def _compute_child_linkage(self, product: str) -> list[dict]:
        records = []
        for other in self.all_products:
            if self.product_types.get(other) != "母基金":
                continue
            other_df = pd.DataFrame(self.edit_cache.get(other, []))
            if other_df.empty:
                continue
            for _, row in other_df.iterrows():
                related = str(row.get("关联产品", "")).strip()
                if related != product and self.reverse_mapping.get(related, related) != product:
                    continue
                inflow = float(row.get("现金流入", 0)) if pd.notna(row.get("现金流入", 0)) else 0
                outflow = float(row.get("现金流出", 0)) if pd.notna(row.get("现金流出", 0)) else 0
                if outflow > 0:
                    records.extend(
                        handle_single_child_linkage(
                            other, row.get("日期"), product, "对子基金申购", outflow,
                            self.sheets_data, self.df_elements, preview_mode=True,
                        )
                    )
                elif inflow > 0:
                    records.extend(
                        handle_single_child_linkage(
                            other, row.get("日期"), product, "对子基金赎回", inflow,
                            self.sheets_data, self.df_elements, preview_mode=True,
                        )
                    )
        return records


STATE = AppState()


# ============================================================
# 可复用的 AG Grid 构建函数
# ============================================================
def build_grid_options(product: str) -> dict:
    display_products = [""] + [STATE.reverse_mapping.get(p, p) for p in STATE.all_products]

    column_defs = [
        {"field": "日期", "headerName": "日期", "editable": True, "width": 110,
         "cellEditor": "agDateStringCellEditor"},
        {"field": "现金流类型", "headerName": "现金流类型", "editable": True, "width": 160},
        {"field": "关联产品", "headerName": "关联产品", "editable": True, "width": 150,
         "cellEditor": "agSelectCellEditor",
         "cellEditorParams": {"values": sorted(display_products, key=lambda x: (x != "", x))}},
        {"field": "现金流入", "headerName": "现金流入", "editable": True, "width": 130,
         "type": "numericColumn", "cellEditor": "agNumberCellEditor"},
        {"field": "现金流出", "headerName": "现金流出", "editable": True, "width": 130,
         "type": "numericColumn", "cellEditor": "agNumberCellEditor"},
        {"field": "期末余额", "headerName": "期末余额", "editable": False, "width": 150,
         "type": "numericColumn",
         "cellClassRules": {"text-red-600": "params.value < 0"}},
        {"field": "备注", "headerName": "备注", "editable": True, "width": 200},
    ]

    return {
        "columnDefs": column_defs,
        "rowData": STATE.get_grid_rows(product),
        "defaultColDef": {"resizable": True, "sortable": True,
            "cellStyle": {"fontSize": "12px"}},
        "rowHeight": 32,
        "rowSelection": "multiple",
        "stopEditingWhenCellsLoseFocus": True,
        "singleClickEdit": False,
        "enableCellChangeFlash": True,
        "undoRedoCellEditing": True,
        "undoRedoCellEditingLimit": 20,
    }


# ============================================================
# 业务辅助
# ============================================================
def recalc_balance(rows: list[dict]):
    """对 rows 原地重新计算期末余额"""
    if not rows:
        return
    first = rows[0].get("期末余额", 0) or 0
    cum = first
    rows[0]["期末余额"] = round(cum, 2)
    for i in range(1, len(rows)):
        inflow = float(rows[i].get("现金流入", 0) or 0)
        outflow = float(rows[i].get("现金流出", 0) or 0)
        cum = cum + inflow - outflow
        rows[i]["期末余额"] = round(cum, 2)


# ============================================================
# 页面: 流动性管理
# ============================================================
class LiquidityPage:
    """流动性管理页面 — 每次通过 install() 重建整个 UI"""

    def __init__(self, parent_container: ui.element):
        self.container = parent_container

    def install(self):
        self.container.clear()
        with self.container:
            self._build()

    def _build(self):
        if not STATE.selected_product:
            ui.label("请从侧边栏选择一个产品").classes("text-gray-500 mt-2")
            return

        product = STATE.selected_product

        # --- 产品信息行 ---
        with ui.row().classes("w-full items-center gap-2"):
            pt = STATE.product_types.get(product, "子基金")
            color = "blue" if pt == "母基金" else "green"
            ui.badge(pt, color=color).props("outline dense")

            # 负余额警告
            rows = STATE.get_grid_rows(product)
            for r in rows:
                if r.get("期末余额", 0) < 0:
                    ft = r.get("现金流类型", "") or "（空）"
                    ui.badge(f"负余额: {r['期末余额']:,.0f} ({ft})", color="red").props("outline dense")
                    break

            ui.space()
            ui.button("参数", on_click=lambda: open_params_dialog(product)).props("flat dense size=sm")

        # --- AG Grid ---
        grid = ui.aggrid(options=build_grid_options(product)).classes("w-full h-[420px] mt-1")

        def on_cell_edit(e):
            if not STATE.selected_product:
                return
            p = STATE.selected_product
            rows = STATE.get_grid_rows(p)
            recalc_balance(rows)
            self.install()

        grid.on("cellValueChanged", on_cell_edit)

        # --- 底部工具栏: 编辑 + 统计 + 保存 ---
        with ui.row().classes("w-full items-center gap-2 mt-1"):
            ui.button("+ 插入行", on_click=lambda: self._insert_row()).props("flat dense size=sm")
            ui.button("+ 复制末行", on_click=lambda: self._copy_last_row()).props("flat dense size=sm")

            rows = STATE.get_grid_rows(product)
            total_in = sum(r.get("现金流入", 0) or 0 for r in rows)
            total_out = sum(r.get("现金流出", 0) or 0 for r in rows)
            final_bal = rows[-1].get("期末余额", 0) if rows else 0
            bal_class = "text-red-600 font-bold" if final_bal < 0 else ""

            ui.label(
                f"共 {len(rows)} 行 | 流入 ¥{total_in:,.2f} | 流出 ¥{total_out:,.2f}"
            ).classes("text-xs text-gray-500")
            ui.label(f"期末余额 ¥{final_bal:,.2f}").classes(f"text-xs {bal_class}")

            ui.space()

            linkage = STATE.compute_linkage(product)
            if linkage:
                ui.badge(f"{len(linkage)} 笔联动", color="orange").props("outline dense")

            ui.button("保存到 Excel", on_click=lambda: open_save_dialog(product), color="primary").props("size=sm")

        # --- 联动预览表 ---
        if linkage:
            with ui.expansion(f"联动预览 ({len(linkage)} 笔)", value=False).classes("w-full mt-1"):
                self._render_linkage_table(linkage)

    def _switch_product(self, new_product: str):
        STATE.selected_product = new_product
        self.install()

    def _insert_row(self):
        p = STATE.selected_product
        if not p:
            return
        rows = STATE.get_grid_rows(p)
        rows.append({col: "" for col in TARGET_COLS})
        rows[-1]["现金流入"] = 0.0
        rows[-1]["现金流出"] = 0.0
        rows[-1]["期末余额"] = 0.0
        STATE.set_grid_rows(p, rows)
        self.install()

    def _copy_last_row(self):
        p = STATE.selected_product
        if not p:
            return
        rows = STATE.get_grid_rows(p)
        if rows:
            new_row = dict(rows[-1])
            new_row["日期"] = ""
            new_row["期末余额"] = 0.0
            rows.append(new_row)
            STATE.set_grid_rows(p, rows)
            recalc_balance(rows)
            self.install()

    def _render_linkage_table(self, linkage: list[dict]):
        columns = [
            {"name": "子基金", "label": "子基金", "field": "子基金"},
            {"name": "母基金操作日", "label": "操作日", "field": "母基金操作日"},
            {"name": "到账日期", "label": "到账日", "field": "到账日期"},
            {"name": "天数", "label": "天数", "field": "到账天数"},
            {"name": "比例", "label": "比例", "field": "比例"},
            {"name": "金额", "label": "金额", "field": "金额"},
            {"name": "子基金操作", "label": "操作", "field": "子基金操作"},
            {"name": "变动后余额", "label": "变动后余额", "field": "变动后余额"},
        ]
        # 格式化
        display_rows = []
        for i, item in enumerate(linkage):
            r = dict(item)
            r["id"] = i
            for k in ("金额", "当前余额", "变动后余额"):
                if k in r and isinstance(r[k], (int, float)):
                    r[k] = f"{r[k]:,.2f}"
            display_rows.append(r)

        ui.table(columns=columns, rows=display_rows, row_key="id").classes("w-full text-sm").props("dense")


# ============================================================
# 页面: 未来待办
# ============================================================
class UpcomingPage:
    def __init__(self, parent_container: ui.element):
        self.container = parent_container

    def install(self):
        self.container.clear()
        with self.container:
            self._build()

    def _build(self):
        events = get_upcoming_events(STATE.sheets_data, STATE.df_elements, STATE.name_mapping)

        if not events:
            with ui.card().classes("w-full bg-gray-50 q-pa-sm"):
                ui.label("未来5日待办 · 0笔").classes("text-gray-500 text-sm")
            return

        today = datetime.now().date()
        by_date: dict[date, list] = {}
        for e in events:
            by_date.setdefault(e["提交日期"], []).append(e)

        today_events = by_date.get(today, [])
        total = len(events)
        buy_count = sum(1 for e in events if e["操作类型"] == "申购")
        sell_count = sum(1 for e in events if e["操作类型"] == "赎回")

        with ui.card().classes("w-full bg-gray-50 q-pa-sm"):
            with ui.row().classes("items-center gap-2"):
                ui.label(f"未来5日待办 · {total}笔 · 申购{buy_count} 赎回{sell_count}").classes("text-sm")
                if today_events:
                    ui.badge(f"今天 {len(today_events)} 笔", color="orange").props("dense")

        with ui.expansion("详情", value=len(today_events) > 0).classes("w-full mt-1"):
            for event_date in sorted(by_date.keys()):
                day_events = by_date[event_date]
                date_str = event_date.strftime("%m-%d")
                weekday_str = ["一","二","三","四","五","六","日"][event_date.weekday()]
                is_today = event_date == today
                tag = " (今天)" if is_today else ""
                items = []
                for e in day_events:
                    op = "申购" if e["操作类型"] == "申购" else "赎回"
                    arr = e["到账日期"].strftime("%m-%d")
                    items.append(f"{op} {e['产品']} {e['金额']:,.0f}元（到账{arr}）")
                ui.label(f"{date_str}周{weekday_str}{tag}  {' | '.join(items)}").classes("text-xs")


# ============================================================
# 对话框: 产品参数编辑
# ============================================================
def open_params_dialog(product: str):
    df = STATE.df_elements
    row_idx = None
    if not df.empty and "产品名称" in df.columns:
        alias_col = _find_alias_column(df)
        if product in df["产品名称"].values:
            row_idx = df[df["产品名称"] == product].index[0]
        elif alias_col and product in df[alias_col].values:
            row_idx = df[df[alias_col] == product].index[0]
        elif product in STATE.name_mapping:
            full = STATE.name_mapping[product]
            if full and full in df["产品名称"].values:
                row_idx = df[df["产品名称"] == full].index[0]

    dlg = ui.dialog()
    with dlg, ui.card().classes("w-[460px]"):
        ui.label(f"{product} - 产品参数").classes("text-base font-bold mb-2")

        if row_idx is not None:
            info = df.loc[row_idx]
            ch = ui.input("申赎渠道", value=str(info.get("申赎渠道", ""))).classes("w-full")
            pd_ = ui.number("申购到账时间(T+N)", value=int(info.get("申购到账时间(T+N)", 0)),
                             min=0, max=30).classes("w-full")
            rd_ = ui.number("赎回到账时间(T+N)", value=int(info.get("赎回到账时间(T+N)", 0)),
                             min=0, max=30).classes("w-full")
            nt = ui.input("备注", value=str(info.get("备注", ""))).classes("w-full")

            def save():
                df.loc[row_idx, "申赎渠道"] = ch.value
                df.loc[row_idx, "申购到账时间(T+N)"] = pd_.value
                df.loc[row_idx, "赎回到账时间(T+N)"] = rd_.value
                df.loc[row_idx, "备注"] = nt.value
                if save_product_elements(df):
                    STATE.df_elements = load_product_elements()
                    ui.notify("参数保存成功", type="positive")
                    dlg.close()
        else:
            ui.label("未找到产品参数，请添加").classes("text-gray-500 mb-3")
            ch = ui.input("申赎渠道", placeholder="如：直销").classes("w-full")
            pd_ = ui.number("申购到账时间(T+N)", min=0, max=30, value=0).classes("w-full")
            rd_ = ui.number("赎回到账时间(T+N)", min=0, max=30, value=0).classes("w-full")
            nt = ui.input("备注").classes("w-full")

            def save():
                new_row = pd.DataFrame([{
                    "产品名称": product, "申赎渠道": ch.value,
                    "申购到账时间(T+N)": pd_.value, "赎回到账时间(T+N)": rd_.value,
                    "备注": nt.value,
                }])
                updated = pd.concat([df, new_row], ignore_index=True)
                if save_product_elements(updated):
                    STATE.df_elements = load_product_elements()
                    ui.notify("参数添加成功", type="positive")
                    dlg.close()

        with ui.row().classes("gap-2 mt-3"):
            ui.button("保存", on_click=save, color="primary")
            ui.button("取消", on_click=dlg.close)

    dlg.open()


# ============================================================
# 对话框: 保存确认
# ============================================================
def open_save_dialog(product: str):
    linkage = STATE.compute_linkage(product)
    dlg = ui.dialog()

    with dlg, ui.card().classes("w-[640px] max-h-[80vh] overflow-auto"):
        ui.label("确认将更改同步到 Excel？").classes("text-base font-bold mb-2")

        if linkage:
            ui.label(f"将同时联动 {len(linkage)} 个子基金").classes("text-green-600 mb-2")
            columns = [
                {"name": "子基金", "label": "子基金", "field": "子基金"},
                {"name": "到账日期", "label": "到账日", "field": "到账日期"},
                {"name": "金额", "label": "金额", "field": "金额"},
                {"name": "子基金操作", "label": "操作", "field": "子基金操作"},
                {"name": "变动后余额", "label": "变动后余额", "field": "变动后余额"},
            ]
            disp_rows = []
            for i, item in enumerate(linkage):
                r = dict(item)
                r["id"] = i
                for k in ("金额", "变动后余额"):
                    if k in r and isinstance(r[k], (int, float)):
                        r[k] = f"{r[k]:,.2f}"
                disp_rows.append(r)
            ui.table(columns=columns, rows=disp_rows, row_key="id").classes("w-full").props("dense")

        with ui.row().classes("gap-2 mt-3"):
            def confirm():
                try:
                    rows = STATE.get_grid_rows(product)
                    edited_df = pd.DataFrame(rows)
                    for col in ["现金流入", "现金流出", "期末余额"]:
                        edited_df[col] = pd.to_numeric(edited_df[col], errors="coerce").fillna(0.0)
                    if "日期" in edited_df.columns:
                        edited_df["日期"] = pd.to_datetime(edited_df["日期"], errors="coerce")

                    df_full = STATE.full_original.get(product, STATE.sheets_data[product]).copy()
                    orig_slice = STATE.original_slice.get(product, edited_df)
                    full_start = len(df_full) - len(orig_slice)
                    if full_start >= 0:
                        df_result = pd.concat(
                            [df_full.iloc[:full_start], edited_df], ignore_index=True
                        )
                    else:
                        df_result = edited_df

                    save_sheet_to_excel(product, df_full, df_result, DEFAULT_FILE)

                    for link in linkage:
                        child = link["子基金"]
                        if child in STATE.sheets_data:
                            child_orig = STATE.sheets_data[child].copy()
                            arr_date = pd.to_datetime(link["到账日期"])
                            child_in = link["金额"] if "申购" in link.get("子基金操作", "") else 0
                            child_out = link["金额"] if "赎回" in link.get("子基金操作", "") else 0
                            _add_cashflow_record(
                                STATE.sheets_data, child, arr_date,
                                link["子基金操作"], product,
                                child_in, child_out, link["变动后余额"],
                            )
                            save_sheet_to_excel(child, child_orig, STATE.sheets_data[child], DEFAULT_FILE)

                    STATE.reload()
                    ui.notify("保存成功", type="positive")
                    dlg.close()
                    # 刷新页面
                    install_all()

                except Exception as e:
                    ui.notify(f"保存失败: {e}", type="negative")

            ui.button("确认保存", on_click=confirm, color="primary")
            ui.button("取消", on_click=dlg.close)

    dlg.open()


# ============================================================
# 应用组装
# ============================================================
main_area: Optional[ui.element] = None
liquidity_page: Optional[LiquidityPage] = None
upcoming_page: Optional[UpcomingPage] = None


def install_all():
    """根据当前页面状态，安装对应页面到 main_area"""
    global liquidity_page, upcoming_page
    if main_area is None:
        return
    if STATE.current_page == "未来待办":
        upcoming_page.install()
    else:
        liquidity_page.install()


@ui.page("/")
def index():
    global main_area, liquidity_page, upcoming_page

    if not os.path.exists(DEFAULT_FILE):
        with ui.column().classes("p-8"):
            ui.label(f"找不到文件: {DEFAULT_FILE}").classes("text-red-500 text-lg")
        return

    if not STATE.sheets_data:
        if not STATE.load(DEFAULT_FILE):
            with ui.column().classes("p-8"):
                ui.label("数据加载失败").classes("text-red-500 text-lg")
            return

    ui.page_title("产品流动性管理系统")

    # --- 左侧导航 ---
    def switch_page(name: str):
        STATE.current_page = name
        install_all()

    def on_sidebar_product_change(product_id: str):
        STATE.selected_product = product_id
        if STATE.current_page != "流动性管理":
            STATE.current_page = "流动性管理"
        install_all()

    with ui.left_drawer(value=True, elevated=True).classes("bg-gray-50") as drawer:
        ui.label("导航").classes("text-base font-bold mb-2")

        # --- 未来待办 ---
        ui.button("未来待办", on_click=lambda: switch_page("未来待办")).props(
            "flat dense align=left"
        ).classes("w-full")

        # --- 流动性管理 (可折叠产品列表) ---
        with ui.expansion("流动性管理", value=True).classes("w-full font-medium"):
            if STATE.all_products:
                for p in STATE.all_products:
                    display = STATE.get_display_name(p)
                    ui.button(
                        display,
                        on_click=lambda _, pid=p: on_sidebar_product_change(pid),
                    ).props("flat dense align=left").classes("w-full pl-4 text-sm")

        ui.separator().classes("my-2")
        if STATE.all_products:
            ui.label(f"共 {len(STATE.all_products)} 个产品").classes("text-xs text-gray-500")

        def reset():
            STATE.reload()
            ui.notify("已重置", type="positive")
            install_all()

        ui.button("重置缓存", on_click=reset).props("flat dense").classes("w-full mt-1 text-xs text-gray-400")

    # --- 顶部 ---
    with ui.header(elevated=True).classes("bg-white text-black q-pa-sm"):
        ui.label("产品流动性管理系统").classes("text-lg font-bold")

    # --- 主内容 ---
    main_area = ui.column().classes("w-full p-2")
    liquidity_page = LiquidityPage(main_area)
    upcoming_page = UpcomingPage(main_area)
    install_all()

    # --- 页脚 ---
    with ui.footer().classes("bg-gray-100 text-gray-400 text-xs q-pa-xs"):
        ui.label("Ctrl+Z 撤销编辑 | 修改母基金现金流自动预览子基金联动")


# ============================================================
# 启动
# ============================================================
if __name__ in {"__main__", "__mp_main__"}:
    ui.run(
        title="产品流动性管理系统",
        host="127.0.0.1",
        port=8080,
        reload=False,
        show=False,
    )

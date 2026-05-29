import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date, timedelta
import os

st.set_page_config(
    layout="wide",
    page_title="产品流动性管理系统",
)

# ----------------------
# 产品参数加载函数
# ----------------------
def load_product_elements():
    """加载产品要素表"""
    file_path = "产品要素表.xlsx"
    
    if os.path.exists(file_path):
        try:
            df = pd.read_excel(file_path)
            # 确保所有列都是正确的数据类型
            if "备注" in df.columns:
                df["备注"] = df["备注"].fillna("").astype(str)
            if "申赎渠道" in df.columns:
                df["申赎渠道"] = df["申赎渠道"].fillna("").astype(str)
            if "产品名称" in df.columns:
                df["产品名称"] = df["产品名称"].fillna("").astype(str)
            
            # 处理简称列（可能有多种命名方式）
            alias_col = None
            for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
                if col_name in df.columns:
                    alias_col = col_name
                    break
            
            if alias_col:
                df[alias_col] = df[alias_col].fillna("").astype(str)
            
            # 确保数值列是数值类型
            for col in ["申购到账时间(T+N)", "赎回到账时间(T+N)"]:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            return df
        except Exception as e:
            st.error(f"加载产品要素表失败：{e}")
            return pd.DataFrame()
    else:
        # 如果文件不存在，创建空的DataFrame
        return pd.DataFrame(columns=[
            "产品名称", "申赎渠道", "申购到账时间(T+N)", 
            "赎回到账时间(T+N)", "备注"
        ])


def create_product_name_mapping(df_elements):
    """创建产品名称映射：简称 -> 全称"""
    if df_elements.empty or "产品名称" not in df_elements.columns:
        return {}
    
    mapping = {}
    
    # 查找简称列
    alias_col = None
    for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
        if col_name in df_elements.columns:
            alias_col = col_name
            break
    
    if alias_col:
        # 创建简称到全称的映射
        for _, row in df_elements.iterrows():
            full_name = row["产品名称"]
            alias = row[alias_col]
            if alias and str(alias).strip():
                mapping[str(alias).strip()] = full_name
            # 也添加全称到自身的映射
            mapping[full_name] = full_name
    else:
        # 如果没有简称列，只使用产品名称
        for _, row in df_elements.iterrows():
            full_name = row["产品名称"]
            mapping[full_name] = full_name
    
    return mapping


def save_product_elements(df):
    """保存产品要素表"""
    try:
        df.to_excel("产品要素表.xlsx", index=False)
        return True
    except Exception as e:
        st.error(f"保存失败：{e}")
        return False


def sort_cashflow_rows_preserve_blank_positions(df):
    """按日期排序，保留空日期行的原始位置，同日期行保持原有顺序"""
    if '日期' not in df.columns:
        return df

    if not df['日期'].isna().any():
        # 使用原始位置作为第二排序键，保持同日期行原有顺序
        df["_orig_idx"] = range(len(df))
        return df.sort_values(
            by=["日期", "_orig_idx"], ascending=[True, True], na_position='last'
        ).drop(columns=["_orig_idx"]).reset_index(drop=True)

    blank_positions = df[df['日期'].isna()].index.tolist()
    non_blank_df = df[df['日期'].notna()].copy()
    non_blank_df["_orig_idx"] = non_blank_df.index.tolist()
    non_blank_df = non_blank_df.sort_values(
        by=["日期", "_orig_idx"], ascending=[True, True], na_position='last'
    ).drop(columns=["_orig_idx"]).reset_index(drop=True)

    result_rows = []
    non_blank_iter = iter(non_blank_df.to_dict('records'))
    blank_set = set(blank_positions)
    for i in range(len(df)):
        if i in blank_set:
            result_rows.append(df.iloc[i].to_dict())
        else:
            result_rows.append(next(non_blank_iter))

    return pd.DataFrame(result_rows)


def resolve_cutoff_start(df, cutoff_date):
    """找到DataFrame中从cutoff_date开始的行索引。
    如果cutoff_date范围内没有数据，则从最近有日期的行开始。"""
    start_idx = None
    for row_i in range(len(df)):
        if pd.notna(df.iloc[row_i]["日期"]) and df.iloc[row_i]["日期"] >= cutoff_date:
            start_idx = row_i
            break
    if start_idx is None:
        for row_i in range(len(df) - 1, -1, -1):
            if pd.notna(df.iloc[row_i]["日期"]):
                start_idx = row_i
                break
    return start_idx


def calc_balance_preserve_first(df):
    """计算期末余额：首行保留Excel原始值，后续行基于首行累加现金流差额"""
    if "期末余额" not in df.columns or len(df) == 0:
        return df
    first_balance = df["期末余额"].iloc[0]
    if pd.isna(first_balance):
        first_balance = 0.0
    else:
        first_balance = float(first_balance)
    net_flow = df["现金流入"].fillna(0) - df["现金流出"].fillna(0)
    df["期末余额"] = first_balance + net_flow.cumsum() - net_flow.iloc[0]
    df["期末余额"] = df["期末余额"].round(2)
    return df


# ============================================================
# Fragment: renders ONLY st.data_editor + balance recalculation.
# Toolbar and save button are outside the fragment so they don't steal focus.
# ============================================================
@st.fragment
def render_product_table_fragment(selected_product, df_processed, display_options_full):
    """Fragment for st.data_editor. Re-runs independently on cell edits."""

    edited_key = f"edited_data_{selected_product}"
    _from_session = edited_key in st.session_state and st.session_state[edited_key] is not None

    if _from_session:
        df = st.session_state[edited_key].copy()
        if "日期" in df.columns:
            df["日期"] = pd.to_datetime(df["日期"], errors='coerce')
        if "现金流入" in df.columns:
            df["现金流入"] = pd.to_numeric(df["现金流入"], errors='coerce').fillna(0.0)
        if "现金流出" in df.columns:
            df["现金流出"] = pd.to_numeric(df["现金流出"], errors='coerce').fillna(0.0)
        if "期末余额" in df.columns:
            df["期末余额"] = pd.to_numeric(df["期末余额"], errors='coerce').fillna(0.0)
        if "期末余额" in df.columns:
            calc_balance_preserve_first(df)
        if "_modified" not in df.columns:
            df["_modified"] = False
    else:
        df = df_processed.copy()
        df["_modified"] = False

    # Info message for new/modified records
    has_new_records = "_is_new" in df.columns and df["_is_new"].any()
    modified_count = df["_modified"].sum() if "_modified" in df.columns else 0
    if has_new_records:
        new_count = df["_is_new"].sum()
        st.info(f"💡 表格中有 {new_count} 笔新增或修改的记录（包括联动记录）")
    elif modified_count > 0:
        st.info(f"📝 表格中有 {modified_count} 行金额已修改")

    # Snapshot of current amounts for change detection later
    prev_inflow = df["现金流入"].fillna(0).astype(float).reset_index(drop=True) if "现金流入" in df.columns else None
    prev_outflow = df["现金流出"].fillna(0).astype(float).reset_index(drop=True) if "现金流出" in df.columns else None

    # Prepare display data: remove internal markers, convert dates
    df_display = df.copy()
    columns_to_hide = ['_is_new', '_modified']
    if '_is_preview' in df_display.columns:
        columns_to_hide.append('_is_preview')
    df_for_display = df_display.drop(columns=[col for col in columns_to_hide if col in df_display.columns])

    # Add visual indicator column for modified rows
    if "_modified" in df.columns:
        df_for_display.insert(0, "变更", df["_modified"].apply(lambda x: "⚠" if x else ""))
    else:
        df_for_display.insert(0, "变更", "")

    if '日期' in df_for_display.columns:
        date_col = pd.to_datetime(df_for_display['日期'], errors='coerce')
        df_for_display['日期'] = date_col.apply(lambda x: x.date() if pd.notna(x) else None)

    # Column configuration
    column_config = {
        "变更": st.column_config.TextColumn("变更", disabled=True,
                                        help="⚠ 表示该行金额已被修改"),
        "日期": st.column_config.DateColumn("日期", format="YYYY-MM-DD"),
        "现金流类型": st.column_config.TextColumn("现金流类型"),
        "关联产品": st.column_config.SelectboxColumn(
            "关联产品",
            options=sorted(display_options_full, key=lambda x: (x != "", x)),
            help="从下拉列表选择关联产品"
        ),
        "现金流入": st.column_config.NumberColumn("现金流入", format="%.2f"),
        "现金流出": st.column_config.NumberColumn("现金流出", format="%.2f"),
        "期末余额": st.column_config.NumberColumn("期末余额", format="%.2f", disabled=True),
        "备注": st.column_config.TextColumn("备注"),
    }

    # Render st.data_editor
    edited_df = st.data_editor(
        df_for_display,
        column_config=column_config,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        key=f"data_editor_{selected_product}",
    )

    # Convert types from data_editor output
    if '现金流入' in edited_df.columns:
        edited_df['现金流入'] = pd.to_numeric(edited_df['现金流入'], errors='coerce').fillna(0.0)
    if '现金流出' in edited_df.columns:
        edited_df['现金流出'] = pd.to_numeric(edited_df['现金流出'], errors='coerce').fillna(0.0)
    if '期末余额' in edited_df.columns:
        edited_df['期末余额'] = pd.to_numeric(edited_df['期末余额'], errors='coerce').fillna(0.0)
    if '日期' in edited_df.columns:
        edited_df['日期'] = pd.to_datetime(edited_df['日期'], errors='coerce')

    # Recalculate balance (preserve first row's Excel value)
    if "期末余额" in edited_df.columns:
        calc_balance_preserve_first(edited_df)

    # ---- Per-row modification detection ----
    curr_inflow = edited_df["现金流入"].fillna(0).astype(float).reset_index(drop=True) if "现金流入" in edited_df.columns else None
    curr_outflow = edited_df["现金流出"].fillna(0).astype(float).reset_index(drop=True) if "现金流出" in edited_df.columns else None

    # Preserve previous _modified state, then add newly detected changes
    if "_modified" in df.columns:
        prev_mod = df["_modified"].reset_index(drop=True)
        new_modified = prev_mod.reindex(range(len(edited_df)), fill_value=False).astype(bool).copy()
    else:
        new_modified = pd.Series(False, index=range(len(edited_df)))

    if prev_inflow is not None and curr_inflow is not None:
        common_len = min(len(prev_inflow), len(curr_inflow))
        for i in range(common_len):
            if abs(prev_inflow.iloc[i] - curr_inflow.iloc[i]) > 0.001 or \
               abs(prev_outflow.iloc[i] - curr_outflow.iloc[i]) > 0.001:
                new_modified.iloc[i] = True
        # For rows beyond common_len (new rows added), keep False

    # Remove display-only column from edited_df
    if "变更" in edited_df.columns:
        edited_df = edited_df.drop(columns=["变更"])

    # Attach _modified to edited_df
    edited_df["_modified"] = new_modified.astype(bool).values

    # Detect if user actually changed data (exclude balance and 变更 columns)
    data_changed = False
    if len(df_for_display) != len(edited_df):
        data_changed = True
    else:
        for col in ["现金流入", "现金流出", "现金流类型", "关联产品", "备注"]:
            if col in df_for_display.columns and col in edited_df.columns:
                a = df_for_display[col].fillna(0).reset_index(drop=True)
                b = edited_df[col].fillna(0).reset_index(drop=True)
                if col in ("现金流入", "现金流出"):
                    a, b = a.astype(float), b.astype(float)
                else:
                    a, b = a.astype(str), b.astype(str)
                if not a.equals(b):
                    data_changed = True
                    break

    # Save to session_state
    st.session_state[edited_key] = edited_df

    # Rerun fragment once to show updated balances (only if user edited data)
    if data_changed:
        st.rerun()


# ----------------------
# 主界面：流动性管理表
# ----------------------

# 自定义CSS优化页面布局
st.markdown("""
<style>
    .block-container {
        padding-top: 0.5rem;
        padding-bottom: 0.25rem;
    }
    h3 {
        margin-top: 0.25rem;
        margin-bottom: 0.25rem;
        font-size: 1.1rem;
    }
    h2 {
        margin-top: 0.25rem;
        margin-bottom: 0.25rem;
        padding-bottom: 0;
    }
    /* Streamlit原生header组件 */
    [data-testid="stHeader"] {
        margin-bottom: 0.25rem;
    }
    .stButton > button {
        padding: 0.15rem 0.35rem;
        font-size: 0.78rem;
    }
    .stSelectbox > div > div {
        min-height: 2rem;
    }
    div[data-testid="stDataframe"] {
        font-size: 0.85rem;
    }
    div[data-testid="stVerticalBlock"] {
        gap: 0.1rem;
    }
</style>
""", unsafe_allow_html=True)


# ---------- 未来待办提醒相关函数 ----------
def get_product_submit_days(product_name, df_elements, name_mapping):
    """从产品要素表中获取产品的申购/赎回到账时间(T+N)"""
    subscribe_days = 0
    redeem_days = 0

    if df_elements.empty:
        return subscribe_days, redeem_days

    if "产品名称" not in df_elements.columns:
        return subscribe_days, redeem_days

    # 尝试通过产品名称匹配
    if product_name in df_elements["产品名称"].values:
        product_info = df_elements[df_elements["产品名称"] == product_name].iloc[0]
        subscribe_days = int(product_info.get("申购到账时间(T+N)", 0))
        redeem_days = int(product_info.get("赎回到账时间(T+N)", 0))
    else:
        # 尝试通过简称匹配
        alias_col = None
        for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
            if col_name in df_elements.columns:
                alias_col = col_name
                break
        if alias_col and product_name in df_elements[alias_col].values:
            product_info = df_elements[df_elements[alias_col] == product_name].iloc[0]
            subscribe_days = int(product_info.get("申购到账时间(T+N)", 0))
            redeem_days = int(product_info.get("赎回到账时间(T+N)", 0))
        elif product_name in name_mapping:
            full_name = name_mapping[product_name]
            if full_name in df_elements["产品名称"].values:
                product_info = df_elements[df_elements["产品名称"] == full_name].iloc[0]
                subscribe_days = int(product_info.get("申购到账时间(T+N)", 0))
                redeem_days = int(product_info.get("赎回到账时间(T+N)", 0))

    return subscribe_days, redeem_days


def get_upcoming_events_from_sheets(sheets_data, df_elements, name_mapping, n_days=5):
    """根据各基金现金流记录的到账日期和产品要素表，反推需提交申购/赎回的日期作为待办事项。

    逻辑：某产品在T日有申购款到账（现金流出），该基金申购到账为N个工作日，
    则需在 T-N 个工作日提交申购申请。
    同理，赎回到账（现金流入）需在到账日前N个工作日提交赎回申请。
    """
    today = datetime.now().date()
    end_date = today + timedelta(days=n_days)
    events = []
    for product_name, df in sheets_data.items():
        if df.empty or "日期" not in df.columns:
            continue

        purchase_days, redeem_days = get_product_submit_days(
            product_name, df_elements, name_mapping
        )

        for _, row in df.iterrows():
            row_date = row["日期"]
            if pd.isna(row_date):
                continue
            row_date = row_date.date() if hasattr(row_date, 'date') else row_date

            inflow = float(row.get("现金流入", 0)) if pd.notna(row.get("现金流入", 0)) else 0
            outflow = float(row.get("现金流出", 0)) if pd.notna(row.get("现金流出", 0)) else 0
            if inflow == 0 and outflow == 0:
                continue

            flow_type = str(row.get("现金流类型", "")).strip()
            related = str(row.get("关联产品", "")).strip()

            # 现金流出=申购, 现金流入=赎回
            if outflow > 0:
                days = purchase_days
                action_type = "申购"
                amount = outflow
            else:
                days = redeem_days
                action_type = "赎回"
                amount = inflow

            # 反推提交日期: 到账日 - N个工作日
            ts = pd.Timestamp(row_date)
            if days > 0:
                submit_date = (ts - pd.offsets.BDay(days)).date()
            else:
                submit_date = row_date

            if submit_date < today or submit_date > end_date:
                continue

            events.append({
                "提交日期": submit_date,
                "到账日期": row_date,
                "产品": product_name,
                "操作类型": action_type,
                "关联产品": related,
                "金额": amount,
                "现金流类型": flow_type,
                "到账天数": days,
            })
    events.sort(key=lambda x: (x["提交日期"], x["产品"]))
    return events


def render_upcoming_reminders(sheets_data, df_elements, name_mapping):
    """渲染未来5天待办提醒（根据产品要素表计算提交日期）"""
    events = get_upcoming_events_from_sheets(sheets_data, df_elements, name_mapping)

    if not events:
        st.markdown("""
        <div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:6px;
                    padding:6px 14px; margin:8px 0;">
            📅 <b>未来5日待办</b> · 0笔
        </div>
        """, unsafe_allow_html=True)
        return

    today = datetime.now().date()
    events_by_date = {}
    for e in events:
        d = e["提交日期"]
        events_by_date.setdefault(d, []).append(e)

    today_events = events_by_date.get(today, [])
    today_count = len(today_events)
    total_count = len(events)
    purchase_count = sum(1 for e in events if e["操作类型"] == "申购")
    redeem_count = sum(1 for e in events if e["操作类型"] == "赎回")

    badge_html = f"<span style='background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:10px;font-size:0.85em;'>今天 {today_count} 笔</span>" if today_count > 0 else ""

    st.markdown(f"""
    <div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:6px;
                padding:6px 14px; margin:8px 0;">
        📅 <b>未来5日待办</b> · {total_count}笔 · 📥申购{purchase_count} 📤赎回{redeem_count} {badge_html}
    </div>
    """, unsafe_allow_html=True)

    default_open = today_count > 0
    with st.expander("详情", expanded=default_open):
        for event_date in sorted(events_by_date.keys()):
            day_events = events_by_date[event_date]
            date_str = event_date.strftime("%m-%d")
            weekday_str = ["一","二","三","四","五","六","日"][event_date.weekday()]
            is_today = event_date == today
            tag = " 🔴" if is_today else ""
            items = []
            for e in day_events:
                if e["操作类型"] == "申购":
                    op = "📥"
                else:
                    op = "📤"
                amt = f"{e['金额']:,.0f}元"
                arrival_str = e["到账日期"].strftime("%m-%d")
                items.append(f"{op} {e['产品']}({e['操作类型']}) {amt}（到账{arrival_str}）")
            st.caption(f"**{date_str}周{weekday_str}{tag}**  {' | '.join(items)}")


# ---------- 1. 加载数据 ----------
DEFAULT_FILE = "基金流动性管理总表_test.xlsx"
file_path = DEFAULT_FILE

# 检查文件是否存在
if not os.path.exists(file_path):
    st.error(f"❌ 找不到文件：{file_path}")
    st.stop()

@st.cache_data(ttl=300)  # 缓存5分钟，减少重新加载
def load_all_sheets(file_path):
    """优化版：使用 xls 对象直接读取，避免重复打开文件"""
    # 一次性创建 ExcelFile 对象
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    sheets_data = {}
    
    # 只读取从第3张sheet开始（索引2）的所有sheet
    target_sheets = xls.sheet_names[2:-4] if len(xls.sheet_names) > 6 else xls.sheet_names[2:]  # 跳过前2个和后4个sheet
    
    for sheet_name in target_sheets:
        # 使用 xls.parse() 而不是 pd.read_excel()，避免重复读取文件
        df = xls.parse(sheet_name)
        
        # 取前7列，标准化列名
        target_cols = ["日期", "现金流类型", "关联产品", "现金流入", "现金流出", "期末余额", "备注"]
        if df.shape[1] >= 7:
            flow_df = df.iloc[:, :7].copy()
            flow_df.columns = target_cols
        elif df.shape[1] >= 6:
            flow_df = df.iloc[:, :6].copy()
            flow_df.columns = target_cols[:6]
            flow_df["备注"] = ""
        elif df.shape[1] >= 5:
            flow_df = df.iloc[:, :5].copy()
            flow_df.columns = ["日期", "现金流类型", "现金流入", "现金流出", "期末余额"]
            flow_df.insert(2, "关联产品", "")
            flow_df["备注"] = ""
        else:
            flow_df = df.copy()
            for i in range(7 - df.shape[1]):
                flow_df[f"col_{i + df.shape[1] + 1}"] = ""
            flow_df.columns = target_cols + list(flow_df.columns[7:])
            flow_df = flow_df.iloc[:, :7]

        # 数值转换
        for col in ["现金流入", "现金流出", "期末余额"]:
            flow_df[col] = pd.to_numeric(flow_df[col], errors="coerce")
        # 日期转换：处理混合类型（datetime字符串/对象 + Excel日期序列号）
        def _fix_excel_date(val):
            if pd.isna(val):
                return val
            if hasattr(val, 'strftime'):
                return val  # 已经是 datetime-like 对象
            try:
                num = float(val)
                if 60 < num < 200000:  # Excel日期序列号范围（约1900~2400年）
                    from datetime import datetime as _dt
                    import numpy as _np
                    return _dt(1899, 12, 30) + timedelta(days=int(num))
            except (ValueError, TypeError):
                pass
            return val

        flow_df["日期"] = flow_df["日期"].apply(_fix_excel_date)
        flow_df["日期"] = pd.to_datetime(flow_df["日期"], errors="coerce")

        # 确保字符串列
        for col in ["现金流类型", "关联产品", "备注"]:
            if col in flow_df.columns:
                flow_df[col] = flow_df[col].fillna("").astype(str)
        
        sheets_data[sheet_name] = flow_df
    
    return sheets_data


try:
    sheets_data = load_all_sheets(file_path)
    all_products = list(sheets_data.keys())
except Exception as e:
    st.error(f"读取文件失败：{e}")
    st.stop()

# 加载产品要素表并创建名称映射
df_elements = load_product_elements()
name_mapping = create_product_name_mapping(df_elements)

# ---------- 侧边栏导航 ----------
with st.sidebar:
    st.markdown("## 📋 导航")
    page = st.radio("", ["📊 流动性管理", "📅 未来待办"], label_visibility="collapsed")

if page == "📅 未来待办":
    st.header("📅 未来5日待办提醒")
    render_upcoming_reminders(sheets_data, df_elements, name_mapping)
    st.stop()

st.header("📊 产品流动性管理")

# 创建反向映射：全称 -> 简称（用于显示）
reverse_mapping = {}
if df_elements is not None and not df_elements.empty:
    alias_col = None
    for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
        if col_name in df_elements.columns:
            alias_col = col_name
            break

    if alias_col:
        for _, row in df_elements.iterrows():
            full_name = row["产品名称"]
            alias = row[alias_col]
            if alias and str(alias).strip():
                reverse_mapping[full_name] = str(alias).strip()

# 过滤 all_products：只保留在产品要素表中存在的产品
def filter_products_by_elements(products, df_elements, name_mapping, reverse_mapping):
    if df_elements.empty or "产品名称" not in df_elements.columns:
        return products
    valid_names = set(df_elements["产品名称"].dropna().unique())
    valid_names.update(reverse_mapping.keys())
    return [p for p in products if p in valid_names or name_mapping.get(p) in valid_names]

all_products = filter_products_by_elements(all_products, df_elements, name_mapping, reverse_mapping)

# 检查是否需要重新加载数据（保存后）
if st.session_state.get('reload_data', False):
    load_all_sheets.clear()
    sheets_data = load_all_sheets(file_path)
    all_products = filter_products_by_elements(list(sheets_data.keys()), df_elements, name_mapping, reverse_mapping)
    st.session_state['reload_data'] = False

# ---------- 2. 产品类型配置（存储在 session_state） ----------
if "product_types" not in st.session_state:
    # 初始默认识别：优先从产品要素表读取，其次根据工作表名称关键词
    st.session_state.product_types = {}
    
    # 查找类型列（可能有多种命名方式）
    type_col = None
    if not df_elements.empty:
        for col_name in ["产品类型", "类型", "基金类型"]:
            if col_name in df_elements.columns:
                type_col = col_name
                break
    
    for p in all_products:
        # 1. 优先从产品要素表读取类型
        if type_col and not df_elements.empty:
            # 尝试通过名称匹配
            if p in df_elements["产品名称"].values:
                product_type = df_elements[df_elements["产品名称"] == p].iloc[0][type_col]
                st.session_state.product_types[p] = str(product_type) if pd.notna(product_type) else "子基金"
            # 尝试通过简称匹配
            else:
                alias_col = None
                for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
                    if col_name in df_elements.columns:
                        alias_col = col_name
                        break
                
                if alias_col and p in df_elements[alias_col].values:
                    product_type = df_elements[df_elements[alias_col] == p].iloc[0][type_col]
                    st.session_state.product_types[p] = str(product_type) if pd.notna(product_type) else "子基金"
                else:
                    # 2. 如果产品要素表没有，根据关键词判断
                    if "母基金" in p or "FOF" in p or "FOF" in p.upper():
                        st.session_state.product_types[p] = "母基金"
                    else:
                        st.session_state.product_types[p] = "子基金"
        else:
            # 3. 如果没有产品要素表，根据关键词判断
            if "母基金" in p or "FOF" in p or "FOF" in p.upper():
                st.session_state.product_types[p] = "母基金"
            else:
                st.session_state.product_types[p] = "子基金"

# ---------- 3. 辅助函数：保存单个 sheet 到 Excel ----------
def save_sheet_to_excel_preserve_rows(sheet_name, df_original, df_updated, wb_path):
    """
    保存数据到Excel，保留原有行结构
    - 新增数据：插入新行
    - 删除数据：清空内容但不删除行
    - 修改数据：更新对应行
    
    Args:
        sheet_name: 工作表名称
        df_original: 原始DataFrame（用于对比）
        df_updated: 更新后的DataFrame
        wb_path: Excel文件路径
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]
    
    # 获取原始数据的行数（不包括表头）
    original_row_count = len(df_original)
    updated_row_count = len(df_updated)
    
    # 情况1：更新的行数 <= 原始行数（有删除或不变）
    if updated_row_count <= original_row_count:
        # 清空所有数据行（从第2行开始，第1行是表头）
        for row in range(2, ws.max_row + 1):
            for col in range(1, 8):
                ws.cell(row=row, column=col, value=None)

        # 写入更新后的数据（从第2行开始）
        for r_idx, (_, row_data) in enumerate(df_updated.iterrows(), start=2):
            for c_idx, value in enumerate(row_data[:7], 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # 情况2：更新的行数 > 原始行数（有新增）
    else:
        # 先更新原有的行
        for r_idx, (_, row_data) in enumerate(df_original.iterrows(), start=2):
            for c_idx, value in enumerate(row_data[:7], 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 在末尾插入新行
        new_rows = df_updated.iloc[original_row_count:]
        for _, row_data in new_rows.iterrows():
            ws.append([row_data.iloc[i] if i < len(row_data) else None for i in range(7)])
    
    wb.save(wb_path)


def sync_to_excel_with_confirmation(selected_product, edited_df, linkage_info=None):
    """
    同步数据到Excel，带确认对话框
    
    Args:
        selected_product: 产品名称
        edited_df: 编辑后的DataFrame
        linkage_info: 联动信息列表
    """
    # 使用session_state存储待保存的数据
    st.session_state[f'pending_save_{selected_product}'] = {
        'df': edited_df,
        'linkage_info': linkage_info or []
    }
    
    # 显示确认对话框
    with st.form(key=f'confirm_save_{selected_product}'):
        st.warning("⚠️ 确认将更改同步到Excel文件？")
        
        if linkage_info:
            st.success(f"🔗 将同时联动 {len(linkage_info)} 个子基金")
            linkage_df = pd.DataFrame(linkage_info)
                    
            # 格式化显示
            styled_df = linkage_df.style.format({
                "金额": "{:,.2f}",
                "当前余额": "{:,.2f}",
                "变动后余额": "{:,.2f}"
                # "比例"列已经是字符串格式（如"50.00%"），不需要额外格式化
            })
                    
            st.dataframe(
                styled_df,
                use_container_width=True,
                hide_index=True,
                height=min(250, len(linkage_info) * 35 + 50)
            )
        
        col_confirm1, col_confirm2 = st.columns(2)
        with col_confirm1:
            confirm_btn = st.form_submit_button("✅ 确认保存", type="primary", use_container_width=True)
        with col_confirm2:
            cancel_btn = st.form_submit_button("❌ 取消", use_container_width=True)
        
        if confirm_btn:
            try:
                # 保存到Excel
                original_df = sheets_data[selected_product]
                save_sheet_to_excel_preserve_rows(
                    selected_product, 
                    original_df, 
                    edited_df, 
                    DEFAULT_FILE
                )
                
                # 如果有联动，也保存到子基金
                if linkage_info:
                    st.info(f"🔗 正在处理 {len(linkage_info)} 个子基金的联动保存...")
                    
                    for linkage in linkage_info:
                        child_product = linkage['子基金']
                        if child_product in sheets_data:
                            try:
                                # 获取子基金的最新数据（可能已经被其他操作修改）
                                child_original = sheets_data[child_product].copy()
                                
                                # 重新计算联动，这次实际添加记录
                                # 从联动信息中提取参数
                                mother_operation_date = pd.to_datetime(linkage['母基金操作日'])
                                amount = linkage['金额']
                                
                                # 判断是申购还是赎回
                                if '申购' in linkage.get('子基金操作', ''):
                                    flow_type = "对子基金申购"
                                else:
                                    flow_type = "对子基金赎回"
                                
                                st.info(f"📝 正在为 {child_product} 添加联动记录：{flow_type}，金额 {amount:,.2f}")
                                
                                # 执行实际的联动操作（preview_mode=False）
                                # 使用单个子基金联动函数
                                handle_single_child_linkage(
                                    selected_product,
                                    mother_operation_date,
                                    child_product,
                                    flow_type,
                                    amount,
                                    preview_mode=False
                                )
                                
                                # 保存更新后的子基金数据
                                child_updated = sheets_data[child_product]
                                save_sheet_to_excel_preserve_rows(
                                    child_product,
                                    child_original,
                                    child_updated,
                                    DEFAULT_FILE
                                )
                                
                                st.success(f"✅ {child_product} 联动记录已保存")
                            except Exception as e:
                                st.error(f"❌ 保存 {child_product} 联动记录失败：{e}")
                                import traceback
                                st.error(traceback.format_exc())
                
                # 清除待保存状态
                if f'pending_save_{selected_product}' in st.session_state:
                    del st.session_state[f'pending_save_{selected_product}']
                
                st.success("✅ 保存成功！母子基金数据已同步更新。")
                
                # 设置标志，触发重新加载数据
                st.session_state['reload_data'] = True
                
                st.cache_data.clear()
                st.rerun()
                
            except Exception as e:
                st.error(f"❌ 保存失败：{e}")
                import traceback
                st.error(traceback.format_exc())
        
        if cancel_btn:
            # 清除待保存状态
            if f'pending_save_{selected_product}' in st.session_state:
                del st.session_state[f'pending_save_{selected_product}']
            st.info("已取消保存")
            st.rerun()


def add_cashflow_record(product_name, date_val, flow_type, related_product, inflow, outflow, balance_after):
    """向指定产品的现金流表追加一行，并返回新的 DataFrame"""
    df = sheets_data[product_name].copy()
    new_row = pd.DataFrame({
        "日期": [date_val],
        "现金流类型": [flow_type],
        "关联产品": [related_product],
        "现金流入": [inflow],
        "现金流出": [outflow],
        "期末余额": [balance_after]
    })
    df = pd.concat([df, new_row], ignore_index=True)
    # 确保数值类型
    for col in ["现金流入", "现金流出", "期末余额"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    sheets_data[product_name] = df
    return df


def get_mother_fund_holdings(mother_fund_name):
    """
    获取母基金持有的子基金份额信息
    
    Args:
        mother_fund_name: 母基金名称
    
    Returns:
        dict: {子基金名称: 持有份额}
    """
    holdings = {}
    
    if mother_fund_name not in sheets_data:
        return holdings
    
    df = sheets_data[mother_fund_name]
    
    # 查找所有关联产品不为空的记录
    related_records = df[df["关联产品"].notna() & (df["关联产品"] != "")].copy()
    
    if related_records.empty:
        return holdings
    
    # 按关联产品分组，获取每个子基金的最新持仓份额
    # 首先按日期排序，确保获取最新的记录
    related_records = related_records.sort_values(by="日期", ascending=False)
    
    # 对每个子基金，只保留最新的记录
    seen_products = set()
    for _, row in related_records.iterrows():
        child_product = str(row["关联产品"]).strip()
        if child_product and child_product in all_products:
            # 如果还没处理过这个子基金，使用其最新余额
            if child_product not in seen_products:
                balance = float(row["期末余额"]) if pd.notna(row["期末余额"]) else 0
                holdings[child_product] = balance
                seen_products.add(child_product)
    
    return holdings


def get_child_product_arrival_days(child_product_name):
    """
    获取子基金的申购/赎回到账时间(T+N)
    
    Args:
        child_product_name: 子基金名称
    
    Returns:
        tuple: (申购到账天数, 赎回到账天数)
    """
    subscribe_days = 0
    redeem_days = 0
    
    # 从产品要素表中查找
    if not df_elements.empty:
        # 尝试通过产品名称匹配
        if child_product_name in df_elements["产品名称"].values:
            product_info = df_elements[df_elements["产品名称"] == child_product_name].iloc[0]
            subscribe_days = int(product_info.get("申购到账时间(T+N)", 0))
            redeem_days = int(product_info.get("赎回到账时间(T+N)", 0))
        else:
            # 尝试通过简称匹配
            alias_col = None
            for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
                if col_name in df_elements.columns:
                    alias_col = col_name
                    break
            
            if alias_col and child_product_name in df_elements[alias_col].values:
                product_info = df_elements[df_elements[alias_col] == child_product_name].iloc[0]
                subscribe_days = int(product_info.get("申购到账时间(T+N)", 0))
                redeem_days = int(product_info.get("赎回到账时间(T+N)", 0))
    
    return subscribe_days, redeem_days


def handle_single_child_linkage(mother_fund_name, date_val, child_product_name, flow_type, amount, preview_mode=True):
    """
    处理母基金对单个指定子基金的申购/赎回联动
    
    联动规则：
    - 母基金申购子基金：母基金流出，指定子基金在T+申购到账日流入
    - 母基金赎回子基金：母基金流入，指定子基金在T+赎回到账日流出
    
    Args:
        mother_fund_name: 母基金名称
        date_val: 交易日期（母基金的操作日期）
        child_product_name: 指定的子基金名称
        flow_type: 现金流类型（申购/赎回）
        amount: 金额（正数）
        preview_mode: 预览模式，True时只计算不实际添加记录
    
    Returns:
        list: 生成的联动记录列表（只有一个子基金）
    """
    linkage_records = []
    
    # 确保 date_val 是 datetime 对象
    if isinstance(date_val, str):
        trade_date = pd.to_datetime(date_val)
    elif isinstance(date_val, pd.Timestamp):
        trade_date = date_val
    else:
        trade_date = pd.to_datetime(date_val)
    
    # 获取子基金的到账时间
    subscribe_days, redeem_days = get_child_product_arrival_days(child_product_name)
    
    # 根据操作类型确定现金流方向和到账日期
    if "申购" in flow_type or "对子基金申购" in flow_type:
        # 母基金申购子基金：
        # - 母基金：现金流出（投资出去）
        # - 子基金：在 T+申购到账日（工作日） 现金流入
        arrival_date = trade_date + pd.offsets.BDay(subscribe_days)

        child_inflow = amount
        child_outflow = 0
        child_flow_type = f"母基金申购到账(T+{subscribe_days})"

    elif "赎回" in flow_type or "对子基金赎回" in flow_type:
        # 母基金赎回子基金：
        # - 母基金：现金流入（收回投资）
        # - 子基金：在 T+赎回到账日（工作日） 现金流出
        arrival_date = trade_date + pd.offsets.BDay(redeem_days)

        child_inflow = 0
        child_outflow = amount
        child_flow_type = f"母基金赎回到账(T+{redeem_days})"
    else:
        return linkage_records
    
    # 获取子基金当前余额
    if child_product_name in sheets_data:
        child_df = sheets_data[child_product_name]
        if not child_df.empty:
            current_balance = float(child_df["期末余额"].iloc[-1]) if pd.notna(child_df["期末余额"].iloc[-1]) else 0
        else:
            current_balance = 0
    else:
        current_balance = 0
    
    # 计算新余额
    new_balance = current_balance + child_inflow - child_outflow
    
    # 如果不是预览模式，才实际添加记录到子基金
    if not preview_mode:
        add_cashflow_record(
            product_name=child_product_name,
            date_val=arrival_date,
            flow_type=child_flow_type,
            related_product=mother_fund_name,  # 关联回母基金
            inflow=child_inflow,
            outflow=child_outflow,
            balance_after=new_balance
        )
    
    trade_date_str = trade_date.strftime("%Y-%m-%d") if pd.notna(trade_date) else ""
    arrival_date_str = arrival_date.strftime("%Y-%m-%d") if pd.notna(arrival_date) else ""

    linkage_records.append({
        "子基金": child_product_name,
        "母基金": mother_fund_name,  # 添加母基金名称
        "母基金操作日": trade_date_str,
        "到账日期": arrival_date_str,
        "到账天数": subscribe_days if "申购" in flow_type else redeem_days,
        "比例": "100.00%",  # 单一子基金，100%
        "金额": amount,
        "子基金操作": child_flow_type,
        "当前余额": current_balance,
        "变动后余额": new_balance
    })
    
    return linkage_records


def handle_mother_fund_cashflow_linkage(mother_fund_name, date_val, flow_type, amount, preview_mode=True):
    """
    处理母基金申购/赎回时的子基金联动
    
    联动规则：
    - 母基金申购子基金：母基金流出，子基金在T+申购到账日流入
    - 母基金赎回子基金：母基金流入，子基金在T+赎回到账日流出
    
    Args:
        mother_fund_name: 母基金名称
        date_val: 交易日期（母基金的操作日期）
        flow_type: 现金流类型（申购/赎回）
        amount: 金额（正数）
        preview_mode: 预览模式，True时只计算不实际添加记录
    
    Returns:
        list: 生成的联动记录列表
    """
    linkage_records = []
    
    # 获取母基金持有的子基金份额
    holdings = get_mother_fund_holdings(mother_fund_name)
    
    if not holdings:
        return linkage_records
    
    # 计算总持有份额
    total_holdings = sum(holdings.values())
    
    if total_holdings == 0:
        return linkage_records
    
    # 确保 date_val 是 datetime 对象
    if isinstance(date_val, str):
        trade_date = pd.to_datetime(date_val)
    elif isinstance(date_val, pd.Timestamp):
        trade_date = date_val
    else:
        trade_date = pd.to_datetime(date_val)
    
    # 按持有比例分配资金到各子基金
    for child_product, child_shares in holdings.items():
        # 计算比例
        ratio = child_shares / total_holdings
        
        # 按比例分配金额
        linked_amount = amount * ratio
        
        # 获取子基金的到账时间
        subscribe_days, redeem_days = get_child_product_arrival_days(child_product)
        
        # 根据操作类型确定现金流方向和到账日期
        if "申购" in flow_type or "对子基金申购" in flow_type:
            # 母基金申购子基金：
            # - 母基金：现金流出（投资出去）
            # - 子基金：在 T+申购到账日 现金流入
            arrival_date = trade_date + pd.offsets.BDay(subscribe_days)

            child_inflow = linked_amount
            child_outflow = 0
            child_flow_type = f"母基金申购到账(T+{subscribe_days})"

        elif "赎回" in flow_type or "对子基金赎回" in flow_type:
            # 母基金赎回子基金：
            # - 母基金：现金流入（收回投资）
            # - 子基金：在 T+赎回到账日（工作日） 现金流出
            arrival_date = trade_date + pd.offsets.BDay(redeem_days)
            
            child_inflow = 0
            child_outflow = linked_amount
            child_flow_type = f"母基金赎回到账(T+{redeem_days})"
        else:
            continue
        
        # 获取子基金当前余额
        if child_product in sheets_data:
            child_df = sheets_data[child_product]
            if not child_df.empty:
                current_balance = float(child_df["期末余额"].iloc[-1]) if pd.notna(child_df["期末余额"].iloc[-1]) else 0
            else:
                current_balance = 0
        else:
            current_balance = 0
        
        # 计算新余额
        new_balance = current_balance + child_inflow - child_outflow
        
        # 如果不是预览模式，才实际添加记录到子基金
        if not preview_mode:
            add_cashflow_record(
                product_name=child_product,
                date_val=arrival_date,
                flow_type=child_flow_type,
                related_product=mother_fund_name,  # 关联回母基金
                inflow=child_inflow,
                outflow=child_outflow,
                balance_after=new_balance
            )
        
        trade_date_str = trade_date.strftime("%Y-%m-%d") if pd.notna(trade_date) else ""
        arrival_date_str = arrival_date.strftime("%Y-%m-%d") if pd.notna(arrival_date) else ""

        linkage_records.append({
            "子基金": child_product,
            "母基金": mother_fund_name,  # 添加母基金名称
            "母基金操作日": trade_date_str,
            "到账日期": arrival_date_str,
            "到账天数": subscribe_days if "申购" in flow_type else redeem_days,
            "比例": f"{ratio:.2%}",
            "金额": linked_amount,
            "子基金操作": child_flow_type,
            "当前余额": current_balance,
            "变动后余额": new_balance
        })
    
    return linkage_records


# ---------- 4. 主界面：产品选择与展示 ----------

# 工具栏行
col_title, col_reset = st.columns([4, 1])
with col_title:
    st.subheader("📊 产品现金流明细")
with col_reset:
    if st.button("🔄 重置缓存", key="reset_all_top", use_container_width=True):
        for p in all_products:
            for k in [f"edited_data_{p}", f"original_data_{p}", f"base_data_{p}"]:
                if k in st.session_state:
                    del st.session_state[k]
        st.success("✅ 已重置")
        st.rerun()

# 创建简称列表用于下拉选择（所有产品共用）
display_options = [""]
for product in all_products:
    display_options.append(reverse_mapping.get(product, product))

if not all_products:
    st.warning("未找到产品数据")
elif st.session_state.get('viewing_product_params', None):
    # ========== 产品参数页面 ==========
    viewing_product = st.session_state['viewing_product_params']

    st.markdown(f"## 📋 {viewing_product} - 产品参数")

    if st.button("⬅️ 返回现金流管理", type="primary"):
        st.session_state['viewing_product_params'] = None
        st.rerun()

    st.markdown("---")

    df_elements_params = load_product_elements()
    product_row = None

    if not df_elements_params.empty:
        alias_col = None
        for col_name in ["简称", "产品简称", "缩写", "产品缩写"]:
            if col_name in df_elements_params.columns:
                alias_col = col_name
                break

        if viewing_product in df_elements_params["产品名称"].values:
            product_row = df_elements_params[df_elements_params["产品名称"] == viewing_product].index[0]
        elif alias_col and viewing_product in df_elements_params[alias_col].values:
            product_row = df_elements_params[df_elements_params[alias_col] == viewing_product].index[0]
        elif name_mapping and viewing_product in name_mapping:
            full_name = name_mapping[viewing_product]
            if full_name in df_elements_params["产品名称"].values:
                product_row = df_elements_params[df_elements_params["产品名称"] == full_name].index[0]

    if product_row is not None:
        product_info = df_elements_params.loc[product_row]

        col_p1, col_p2 = st.columns(2)
        with col_p1:
            new_channel = st.text_input("申赎渠道", value=str(product_info.get("申赎渠道", "")), key=f"edit_channel_{viewing_product}")
            new_purchase_days = st.number_input("申购到账时间(T+N)", min_value=0, max_value=30,
                                                value=int(product_info.get("申购到账时间(T+N)", 0)), key=f"edit_purchase_{viewing_product}")
        with col_p2:
            new_redeem_days = st.number_input("赎回到账时间(T+N)", min_value=0, max_value=30,
                                              value=int(product_info.get("赎回到账时间(T+N)", 0)), key=f"edit_redeem_{viewing_product}")
            new_notes = st.text_input("备注", value=str(product_info.get("备注", "")), key=f"edit_notes_{viewing_product}")

        col_save1, col_save2 = st.columns([1, 3])
        with col_save1:
            if st.button("💾 保存参数", type="primary", key=f"save_product_params_{viewing_product}"):
                df_elements_params.loc[product_row, "申赎渠道"] = new_channel
                df_elements_params.loc[product_row, "申购到账时间(T+N)"] = new_purchase_days
                df_elements_params.loc[product_row, "赎回到账时间(T+N)"] = new_redeem_days
                df_elements_params.loc[product_row, "备注"] = new_notes
                if save_product_elements(df_elements_params):
                    st.success("✅ 产品参数保存成功！")
                    st.rerun()
        with col_save2:
            if st.button("🔄 取消编辑", key=f"cancel_edit_{viewing_product}"):
                st.session_state['viewing_product_params'] = None
                st.rerun()
    else:
        st.info(f"⚠️ 未找到「{viewing_product}」的参数信息，请在产品要素表中添加")

        col_add1, col_add2 = st.columns(2)
        with col_add1:
            add_channel = st.text_input("申赎渠道", placeholder="如：直销、代销", key=f"add_channel_{viewing_product}")
            add_purchase_days = st.number_input("申购到账时间(T+N)", min_value=0, max_value=30, value=0, key=f"add_purchase_{viewing_product}")
        with col_add2:
            add_redeem_days = st.number_input("赎回到账时间(T+N)", min_value=0, max_value=30, value=0, key=f"add_redeem_{viewing_product}")
            add_notes = st.text_input("备注", key=f"add_notes_{viewing_product}")

        if st.button("➕ 添加产品参数", type="primary", key=f"add_product_params_{viewing_product}"):
            new_row = pd.DataFrame({
                "产品名称": [viewing_product],
                "申赎渠道": [add_channel],
                "申购到账时间(T+N)": [add_purchase_days],
                "赎回到账时间(T+N)": [add_redeem_days],
                "备注": [add_notes]
            })
            df_elements_params = pd.concat([df_elements_params, new_row], ignore_index=True)
            if save_product_elements(df_elements_params):
                st.success("✅ 产品参数添加成功！")
                st.rerun()

else:
    # ========== st.tabs 展示所有产品 ==========
    tab_labels = [reverse_mapping.get(p, p) for p in all_products]
    tabs = st.tabs(tab_labels)

    for tab, selected_product in zip(tabs, all_products):
        with tab:
            # ---- 数据加载与处理 ----
            edited_session_key = f"edited_data_{selected_product}"
            original_session_key = f"original_data_{selected_product}"

            _from_session = edited_session_key in st.session_state and st.session_state[edited_session_key] is not None
            if _from_session:
                df = st.session_state[edited_session_key].copy()
                df["日期"] = pd.to_datetime(df["日期"], errors='coerce')
                if '_is_new' in df.columns:
                    df = df.drop(columns=['_is_new'])

                if original_session_key in st.session_state:
                    original_df = st.session_state[original_session_key].copy()
                else:
                    original_df = sheets_data[selected_product].copy()
                    cutoff_date = pd.Timestamp.today() - timedelta(days=5)
                    original_df["日期"] = pd.to_datetime(original_df["日期"], errors='coerce')
                    orig_start_idx = resolve_cutoff_start(original_df, cutoff_date)
                    if orig_start_idx is not None:
                        original_df = original_df.iloc[orig_start_idx:].copy()
                    else:
                        original_df = original_df[original_df["日期"].isna()].copy()
                    original_df = original_df.reset_index(drop=True)
                    original_df["现金流入"] = original_df["现金流入"].fillna(0).astype(float).round(2)
                    original_df["现金流出"] = original_df["现金流出"].fillna(0).astype(float).round(2)
                    original_df["期末余额"] = original_df["期末余额"].fillna(0).astype(float).round(2)
                    if reverse_mapping:
                        original_df["关联产品"] = original_df["关联产品"].apply(
                            lambda x: reverse_mapping.get(str(x).strip(), x) if str(x).strip() else ""
                        )
                    original_df = sort_cashflow_rows_preserve_blank_positions(original_df)
                    st.session_state[original_session_key] = original_df.copy()
            else:
                _from_session = False
                df = sheets_data[selected_product].copy()
                original_df = df.copy()

                cutoff_date = pd.Timestamp.today() - timedelta(days=5)
                original_df["日期"] = pd.to_datetime(original_df["日期"], errors='coerce')
                orig_start_idx = resolve_cutoff_start(original_df, cutoff_date)
                if orig_start_idx is not None:
                    original_df = original_df.iloc[orig_start_idx:].copy()
                else:
                    original_df = original_df[original_df["日期"].isna()].copy()
                original_df = original_df.reset_index(drop=True)
                original_df["现金流入"] = original_df["现金流入"].fillna(0).astype(float).round(2)
                original_df["现金流出"] = original_df["现金流出"].fillna(0).astype(float).round(2)
                original_df["期末余额"] = original_df["期末余额"].fillna(0).astype(float).round(2)
                if reverse_mapping:
                    original_df["关联产品"] = original_df["关联产品"].apply(
                        lambda x: reverse_mapping.get(str(x).strip(), x) if str(x).strip() else ""
                    )
                original_df = sort_cashflow_rows_preserve_blank_positions(original_df)
                st.session_state[original_session_key] = original_df.copy()

            # 过滤：从today-5开始，没有数据则从最近有日期的行开始
            cutoff_date = pd.Timestamp.today() - timedelta(days=5)
            df["日期"] = pd.to_datetime(df["日期"], errors='coerce')

            start_idx = resolve_cutoff_start(df, cutoff_date)

            if start_idx is not None:
                if _from_session:
                    nat_before = df.iloc[:start_idx]["日期"].isna()
                    rows_to_keep = list(range(start_idx, len(df)))
                    rows_to_keep += nat_before[nat_before].index.tolist()
                    rows_to_keep = sorted(set(rows_to_keep))
                    df = df.iloc[rows_to_keep].copy()
                else:
                    df = df.iloc[start_idx:].copy()
            else:
                df = df[df["日期"].isna()].copy()

            df = df.reset_index(drop=True)
            df["现金流入"] = df["现金流入"].fillna(0).astype(float).round(2)
            df["现金流出"] = df["现金流出"].fillna(0).astype(float).round(2)
            df["期末余额"] = df["期末余额"].fillna(0).astype(float).round(2)
            df = sort_cashflow_rows_preserve_blank_positions(df)

            if reverse_mapping:
                df["关联产品"] = df["关联产品"].apply(
                    lambda x: reverse_mapping.get(str(x).strip(), x) if str(x).strip() else ""
                )

            # ---- 联动检测 ----
            linkage_info = []

            if st.session_state.product_types.get(selected_product) == "母基金":
                if "日期" in df.columns:
                    df["日期"] = pd.to_datetime(df["日期"], errors='coerce')
                    original_df["日期"] = pd.to_datetime(original_df["日期"], errors='coerce')

                new_row_indices = []
                for row_i in range(len(df)):
                    if row_i >= len(original_df):
                        new_row_indices.append(row_i)
                    else:
                        orig_row = original_df.iloc[row_i]
                        edit_row = df.iloc[row_i]
                        orig_related = str(orig_row.get("关联产品", "")).strip()
                        edit_related = str(edit_row.get("关联产品", "")).strip()
                        orig_inflow = round(float(orig_row.get("现金流入", 0)) if pd.notna(orig_row.get("现金流入", 0)) else 0, 2)
                        edit_inflow = round(float(edit_row.get("现金流入", 0)) if pd.notna(edit_row.get("现金流入", 0)) else 0, 2)
                        orig_outflow = round(float(orig_row.get("现金流出", 0)) if pd.notna(orig_row.get("现金流出", 0)) else 0, 2)
                        edit_outflow = round(float(edit_row.get("现金流出", 0)) if pd.notna(edit_row.get("现金流出", 0)) else 0, 2)
                        if (orig_related != edit_related or
                            abs(orig_inflow - edit_inflow) > 0.01 or
                            abs(orig_outflow - edit_outflow) > 0.01):
                            new_row_indices.append(row_i)

                df["_is_new"] = False
                for row_i in new_row_indices:
                    df.iloc[row_i, df.columns.get_loc("_is_new")] = True

                for row_i in range(len(df)):
                    row = df.iloc[row_i]
                    related_product = str(row.get("关联产品", "")).strip()
                    inflow = float(row.get("现金流入", 0)) if pd.notna(row.get("现金流入", 0)) else 0
                    outflow = float(row.get("现金流出", 0)) if pd.notna(row.get("现金流出", 0)) else 0
                    date_val = row.get("日期")

                    if related_product and related_product in all_products:
                        if outflow > 0:
                            linkage_info.extend(handle_single_child_linkage(
                                selected_product, date_val, related_product, "对子基金申购", outflow, preview_mode=True
                            ))
                        elif inflow > 0:
                            linkage_info.extend(handle_single_child_linkage(
                                selected_product, date_val, related_product, "对子基金赎回", inflow, preview_mode=True
                            ))

            elif st.session_state.product_types.get(selected_product) == "子基金":
                for other_product in all_products:
                    if st.session_state.product_types.get(other_product) == "母基金":
                        try:
                            other_edited_key = f"edited_data_{other_product}"
                            if other_edited_key in st.session_state and st.session_state[other_edited_key] is not None:
                                other_df = st.session_state[other_edited_key].copy()
                            else:
                                other_df = sheets_data[other_product].copy()

                            if other_df is not None and not other_df.empty:
                                for row_i in range(len(other_df)):
                                    row = other_df.iloc[row_i]
                                    related_product = str(row.get("关联产品", "")).strip()

                                    if (related_product == selected_product or
                                        related_product == reverse_mapping.get(selected_product, selected_product) or
                                        name_mapping.get(related_product) == selected_product):
                                        inflow = float(row.get("现金流入", 0)) if pd.notna(row.get("现金流入", 0)) else 0
                                        outflow = float(row.get("现金流出", 0)) if pd.notna(row.get("现金流出", 0)) else 0
                                        date_val = row.get("日期")

                                        if outflow > 0:
                                            linkage_info.extend(handle_single_child_linkage(
                                                other_product, date_val, selected_product, "对子基金申购", outflow, preview_mode=True
                                            ))
                                        elif inflow > 0:
                                            linkage_info.extend(handle_single_child_linkage(
                                                other_product, date_val, selected_product, "对子基金赎回", inflow, preview_mode=True
                                            ))
                        except Exception:
                            pass

            # 子基金联动预览集成
            if st.session_state.product_types.get(selected_product) == "子基金":
                if len(df) > 0:
                    is_linkage_mask = df["现金流类型"].astype(str).str.contains("母基金|申购到账|赎回到账", na=False) if "现金流类型" in df.columns else pd.Series(False, index=df.index)
                    if is_linkage_mask.any():
                        df = df[~is_linkage_mask].copy()
                        df = df.reset_index(drop=True)

                if linkage_info:
                    preview_records = []
                    for linkage in linkage_info:
                        arrival_date = pd.to_datetime(linkage['到账日期'])
                        inflow = linkage['金额'] if '申购' in linkage.get('子基金操作', '') else 0
                        outflow = linkage['金额'] if '赎回' in linkage.get('子基金操作', '') else 0
                        preview_records.append({
                            "日期": arrival_date,
                            "现金流类型": linkage['子基金操作'],
                            "关联产品": linkage.get('母基金', ''),
                            "现金流入": inflow,
                            "现金流出": outflow,
                            "期末余额": None,
                            "_is_new": True
                        })

                    if preview_records:
                        df_with_preview = pd.concat([df, pd.DataFrame(preview_records)], ignore_index=True)
                        df_with_preview = sort_cashflow_rows_preserve_blank_positions(df_with_preview)
                        df = df_with_preview
                        df["现金流入"] = df["现金流入"].fillna(0).astype(float)
                        df["现金流出"] = df["现金流出"].fillna(0).astype(float)
                        df["期末余额"] = df["期末余额"].fillna(0).astype(float)
                        df["_is_new"] = df["_is_new"].fillna(False)

            # 计算余额：首行保留Excel原始值，后续行累加现金流差额
            if "期末余额" in df.columns:
                calc_balance_preserve_first(df)

            # 产品信息行（类型 + 负余额预警 + 编辑参数）
            product_type = st.session_state.product_types.get(selected_product, "子基金")
            col_info, col_warn, col_edit = st.columns([3, 3, 1])
            with col_info:
                st.caption(f"📋 **{selected_product}** | 类型：**{product_type}**")
            with col_warn:
                if "期末余额" in df.columns and (df["期末余额"] < 0).any():
                    first_neg = df[df["期末余额"] < 0].iloc[0]
                    flow_type = str(first_neg.get("现金流类型", "")).strip() or "（空）"
                    bal = float(first_neg["期末余额"])
                    neg_count = (df["期末余额"] < 0).sum()
                    more = f" (+{neg_count - 1}行)" if neg_count > 1 else ""
                    st.caption(f"⚠️ `{flow_type}` → 余额 {bal:,.0f}{more}")
            with col_edit:
                if st.button("✎ 参数", key=f"edit_params_{selected_product}"):
                    st.session_state['viewing_product_params'] = selected_product
                    st.rerun()

            # 渲染表格（fragment: cell edits don't trigger full page reload）
            render_product_table_fragment(selected_product, df, display_options)

            # ---- 工具栏（在fragment外部，避免st.rerun时抢占焦点） ----
            edited_key = f"edited_data_{selected_product}"
            if edited_key in st.session_state and st.session_state[edited_key] is not None:
                edited_df = st.session_state[edited_key]

                col_tool1, col_tool2, col_tool3, col_tool4 = st.columns([1, 1, 1, 3])
                with col_tool1:
                    new_row_count = st.number_input(
                        "行数", min_value=1, max_value=500, value=len(edited_df),
                        key=f"row_num_{selected_product}"
                    )
                with col_tool2:
                    if st.button("➕ 插入行", key=f"insert_{selected_product}"):
                        new_row = pd.DataFrame([{col: None for col in edited_df.columns}])
                        edited_df = pd.concat([edited_df, new_row], ignore_index=True)
                        if "现金流入" in edited_df.columns:
                            edited_df["现金流入"] = edited_df["现金流入"].fillna(0.0)
                        if "现金流出" in edited_df.columns:
                            edited_df["现金流出"] = edited_df["现金流出"].fillna(0.0)
                        if "期末余额" in edited_df.columns:
                            calc_balance_preserve_first(edited_df)
                        st.session_state[edited_key] = edited_df
                        st.rerun()
                with col_tool3:
                    if st.button("📋 复制末行", key=f"dup_{selected_product}"):
                        if len(edited_df) > 0:
                            last_row = edited_df.iloc[[-1]].copy()
                            last_row["日期"] = None
                            last_row["期末余额"] = None
                            edited_df = pd.concat([edited_df, last_row], ignore_index=True)
                            if "期末余额" in edited_df.columns:
                                calc_balance_preserve_first(edited_df)
                            st.session_state[edited_key] = edited_df
                            st.rerun()
                with col_tool4:
                    final_balance = edited_df['期末余额'].iloc[-1] if len(edited_df) > 0 and '期末余额' in edited_df.columns else 0
                    balance_str = f"| 期末余额 ¥{final_balance:,.2f}"
                    if final_balance < 0:
                        balance_str += " ⚠️ 负余额！"
                    st.caption(f"共 {len(edited_df)} 行 | 总流入 ¥{edited_df['现金流入'].sum():,.2f} | 总流出 ¥{edited_df['现金流出'].sum():,.2f} {balance_str}")

                # ---- 保存按钮（在fragment外部） ----
                col_save1, col_save2 = st.columns([1, 5])
                with col_save1:
                    if st.button("💾 保存到Excel", key=f"save_{selected_product}", type="primary"):
                        sync_to_excel_with_confirmation(selected_product, edited_df, linkage_info)

st.markdown("---")
st.caption("💡 提示：点击标签切换产品 | 点击「⚙️ 编辑参数」修改产品参数 | 修改母基金现金流会自动预览子基金联动")